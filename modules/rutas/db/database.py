import logging
import os
from sqlalchemy import ForeignKey, Integer, String, Float, UniqueConstraint, select, delete, text, func
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column, relationship, joinedload
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from modules.camiones.config import settings

logger = logging.getLogger(__name__)

DATABASE_URL = os.getenv("DATABASE_URL") or settings.database_url or "sqlite+aiosqlite:///./auditoria.db"

engine = create_async_engine(
    DATABASE_URL, echo=False,
    connect_args={"check_same_thread": False}
)
async_session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


class Base(DeclarativeBase):
    pass


class RutaMadreDb(Base):
    __tablename__ = "rutas_madre"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    sucursal: Mapped[str] = mapped_column(String(50), nullable=False)
    nombre: Mapped[str] = mapped_column(String(100), nullable=False)

    hijas: Mapped[list["RutaHijaDb"]] = relationship(
        "RutaHijaDb", back_populates="madre", cascade="all, delete-orphan",
        passive_deletes=True
    )

    __table_args__ = (
        UniqueConstraint("sucursal", "nombre", name="uq_madre_sucursal_nombre"),
    )


class RutaHijaDb(Base):
    __tablename__ = "rutas_hijas"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    ruta_madre_id: Mapped[int] = mapped_column(
        Integer, ForeignKey("rutas_madre.id", ondelete="CASCADE"), nullable=False
    )
    ruta_hija: Mapped[str] = mapped_column(String(200), nullable=False)
    flete: Mapped[float] = mapped_column(Float, default=0.0)
    metodo: Mapped[str | None] = mapped_column(String(50), nullable=True)

    madre: Mapped["RutaMadreDb"] = relationship(
        "RutaMadreDb", back_populates="hijas"
    )


async def init_db():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
        await conn.execute(text("PRAGMA foreign_keys = ON"))
    logger.info("Tablas de rutas creadas/verificadas")


def _resolve_seed_path(filename: str = "BBDDs_SL.xlsx") -> str:
    paths = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", filename),
        os.path.join(os.getcwd(), filename),
        filename,
    ]
    for p in paths:
        if os.path.exists(p):
            return p
    return filename


def _parse_float_safe(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


async def obtener_madres_con_hijas(sucursal: str | None = None) -> list[RutaMadreDb]:
    async with async_session_factory() as session:
        stmt = select(RutaMadreDb).options(joinedload(RutaMadreDb.hijas)).order_by(RutaMadreDb.sucursal, RutaMadreDb.nombre)
        if sucursal:
            stmt = stmt.where(RutaMadreDb.sucursal == sucursal)
        result = await session.execute(stmt)
        return list(result.unique().scalars().all())


async def obtener_madre_por_id(madre_id: int) -> RutaMadreDb | None:
    async with async_session_factory() as session:
        stmt = select(RutaMadreDb).options(joinedload(RutaMadreDb.hijas)).where(RutaMadreDb.id == madre_id)
        result = await session.execute(stmt)
        return result.unique().scalar_one_or_none()


async def crear_madre(sucursal: str, nombre: str) -> RutaMadreDb:
    async with async_session_factory() as session:
        madre = RutaMadreDb(sucursal=sucursal, nombre=nombre)
        session.add(madre)
        await session.flush()
        await session.commit()
        stmt = select(RutaMadreDb).options(joinedload(RutaMadreDb.hijas)).where(RutaMadreDb.id == madre.id)
        result = await session.execute(stmt)
        return result.unique().scalar_one()


async def actualizar_madre(madre_id: int, datos: dict) -> RutaMadreDb | None:
    async with async_session_factory() as session:
        madre = await session.get(RutaMadreDb, madre_id)
        if not madre:
            return None
        for key, val in datos.items():
            if val is not None:
                setattr(madre, key, val)
        await session.commit()
        stmt = select(RutaMadreDb).options(joinedload(RutaMadreDb.hijas)).where(RutaMadreDb.id == madre_id)
        result = await session.execute(stmt)
        return result.unique().scalar_one_or_none()


async def eliminar_madre(madre_id: int) -> bool:
    async with async_session_factory() as session:
        madre = await session.get(RutaMadreDb, madre_id)
        if not madre:
            return False
        await session.delete(madre)
        await session.commit()
        return True


async def crear_hija(ruta_madre_id: int, ruta_hija: str, flete: float = 0.0, metodo: str | None = None) -> RutaHijaDb:
    async with async_session_factory() as session:
        hija = RutaHijaDb(ruta_madre_id=ruta_madre_id, ruta_hija=ruta_hija, flete=flete, metodo=metodo)
        session.add(hija)
        await session.commit()
        await session.refresh(hija)
        return hija


async def actualizar_hija(hija_id: int, datos: dict) -> RutaHijaDb | None:
    async with async_session_factory() as session:
        hija = await session.get(RutaHijaDb, hija_id)
        if not hija:
            return None
        for key, val in datos.items():
            setattr(hija, key, val)
        await session.commit()
        await session.refresh(hija)
        return hija


async def eliminar_hija(hija_id: int) -> bool:
    async with async_session_factory() as session:
        hija = await session.get(RutaHijaDb, hija_id)
        if not hija:
            return False
        await session.delete(hija)
        await session.commit()
        return True


async def seed_desde_excel(file_path: str | None = None) -> dict:
    import openpyxl
    path = file_path or _resolve_seed_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    if "RUTAS" not in wb.sheetnames:
        wb.close()
        raise KeyError(f"La hoja 'RUTAS' no existe en {path}. Hojas: {wb.sheetnames}")

    ws = wb["RUTAS"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    madres_map: dict[tuple[str, str], int] = {}
    creadas = 0
    omitidas = 0

    async with async_session_factory() as session:
        for row in rows:
            if len(row) < 5:
                omitidas += 1
                continue
            sucursal = str(row[0]).strip() if row[0] else ""
            madre_nombre = str(row[1]).strip() if row[1] else ""
            ruta_hija_raw = row[2]
            flete_raw = row[3]
            metodo_raw = row[4] if len(row) > 4 else None

            if not sucursal or not madre_nombre or not ruta_hija_raw:
                omitidas += 1
                continue

            ruta_hija = str(ruta_hija_raw).strip()
            flete_val = _parse_float_safe(flete_raw)
            metodo_val = str(metodo_raw).strip() if metodo_raw and str(metodo_raw).strip() else None

            key = (sucursal, madre_nombre)
            if key not in madres_map:
                stmt = select(RutaMadreDb).where(
                    RutaMadreDb.sucursal == key[0], RutaMadreDb.nombre == key[1]
                )
                result = await session.execute(stmt)
                madre = result.scalar_one_or_none()
                if not madre:
                    madre = RutaMadreDb(sucursal=key[0], nombre=key[1])
                    session.add(madre)
                    await session.flush()
                madres_map[key] = madre.id

            # Evitar duplicados: verificar si ya existe (ruta_madre_id, ruta_hija)
            stmt_h = select(RutaHijaDb).where(
                RutaHijaDb.ruta_madre_id == madres_map[key],
                RutaHijaDb.ruta_hija == ruta_hija,
            )
            existing = (await session.execute(stmt_h)).scalar_one_or_none()
            if existing:
                omitidas += 1
                continue

            hija = RutaHijaDb(
                ruta_madre_id=madres_map[key],
                ruta_hija=ruta_hija,
                flete=flete_val,
                metodo=metodo_val,
            )
            session.add(hija)
            creadas += 1

        await session.commit()

    wb.close()
    return {"madres": len(madres_map), "hijas": creadas, "omitidas": omitidas}


async def obtener_promedios_por_madre(sucursal: str | None = None) -> list[dict]:
    """Retorna promedio de flete agrupado por ruta_madre + sucursal."""
    async with async_session_factory() as session:
        stmt = select(
            RutaMadreDb.sucursal,
            RutaMadreDb.nombre,
            func.avg(RutaHijaDb.flete),
            func.count(RutaHijaDb.id),
            func.min(RutaHijaDb.flete),
            func.max(RutaHijaDb.flete),
        ).join(
            RutaHijaDb, RutaMadreDb.id == RutaHijaDb.ruta_madre_id
        ).group_by(RutaMadreDb.sucursal, RutaMadreDb.nombre).order_by(RutaMadreDb.sucursal, RutaMadreDb.nombre)
        if sucursal:
            stmt = stmt.where(RutaMadreDb.sucursal == sucursal)
        result = await session.execute(stmt)
        return [
            {
                "sucursal": row[0],
                "ruta_madre": row[1],
                "promedio_flete": round(row[2] or 0, 0),
                "total_rutas": row[3],
                "min_flete": row[4] or 0,
                "max_flete": row[5] or 0,
            }
            for row in result
        ]


async def obtener_hijas_con_flete(sucursal: str | None = None) -> list[dict]:
    """Retorna todas las rutas_hijas con flete, agrupadas por ruta_madre + sucursal."""
    async with async_session_factory() as session:
        stmt = select(RutaHijaDb).join(RutaHijaDb.madre).options(joinedload(RutaHijaDb.madre))
        if sucursal:
            stmt = stmt.where(RutaMadreDb.sucursal == sucursal)
        stmt = stmt.order_by(RutaMadreDb.sucursal, RutaMadreDb.nombre, RutaHijaDb.ruta_hija)
        result = await session.execute(stmt)
        hijas = result.unique().scalars().all()
        return [
            {
                "id": h.id,
                "ruta_hija": h.ruta_hija,
                "flete": h.flete,
                "metodo": h.metodo,
                "ruta_madre": h.madre.nombre,
                "sucursal": h.madre.sucursal,
            }
            for h in hijas
        ]


async def limpiar_todo():
    async with async_session_factory() as session:
        await session.execute(delete(RutaHijaDb))
        await session.execute(delete(RutaMadreDb))
        await session.commit()
    logger.info("Tablas de rutas limpiadas")


async def obtener_flete_por_ruta_o_madre(ruta: str | None, ruta_madre: str | None) -> float:
    """Busca flete promedio en rutas_hijas usando ruta_madre (o ruta como fallback)."""
    nombre_busqueda = ruta_madre or ruta or ""
    if (ruta or "").strip().lower() == "local":
        nombre_busqueda = "MERCADO"

    if not nombre_busqueda:
        return 0.0

    async with async_session_factory() as session:
        stmt = select(func.avg(RutaHijaDb.flete)).join(RutaHijaDb.madre).where(
            func.upper(RutaMadreDb.nombre) == nombre_busqueda.strip().upper()
        )
        result = await session.execute(stmt)
        flete = result.scalar()
        return round(flete, 2) if flete is not None else 0.0
