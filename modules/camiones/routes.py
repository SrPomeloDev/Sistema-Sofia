"""
routes.py — API REST endpoints para el módulo Camiones.
"""

import logging
import json
import asyncio
from datetime import datetime, timezone
from pathlib import Path

from sqlalchemy import select
from fastapi import APIRouter, HTTPException, status

from modules.camiones.config import settings
from modules.camiones.models import (
    CamionCreate,
    CamionUpdate,
    CamionResponse,
    UpdateSheetResponse,
    AuditEntry,
    SyncStatusResponse,
    FletePromedioResponse,
    PromedioFleteRutaResponse,
    RutaPrecioUpdate,
)
from modules.camiones.db.database import (
    init_db,
    crear_registro_auditoria,
    actualizar_estado_auditoria,
    obtener_historial,
    obtener_todos_camiones,
    obtener_camion_por_placa,
    obtener_camion_por_fila,
    guardar_camiones_bulk,
    upsert_camiones_desde_sheets,
    crear_camion_local,
    actualizar_camion_local,
    eliminar_camion_local,
    obtener_siguiente_nro_sucursal,
    obtener_max_fila_id,
    obtener_pendientes_sincronizacion_count,
    obtener_total_camiones_count,
    obtener_ultimo_cambio,
    obtener_camiones_por_sucursal,
    obtener_promedio_flete_por_sucursal,
    recalcular_promedios_ruta,
    obtener_promedios_ruta,
    guardar_promedio_ruta,
    eliminar_promedio_ruta,
    obtener_todas_tarifas,
    cargar_tarifas_desde_lista,
    obtener_clasificaciones_tarifas,
    eliminar_tarifa,
    seed_rutas_desde_excel,
    obtener_rutas_madres,
    obtener_rutas_hijas,
)
from modules.camiones.services.sheets import sheets_client
from modules.camiones.services.queue import UpdateQueue, QueueItem
from modules.camiones.services.excel_parser import parse_excel_camiones
from modules.camiones.services.bootstrap import bootstrap_sheets
from modules.camiones.auth import verify_credentials, create_session, verify_session, destroy_session

logger = logging.getLogger(__name__)

# Cabeceras por defecto para Google Sheets
HEADERS = [
    "Nº",
    "Nº placa ",
    "Estado de trabajo",
    "Tipo de combustible",
    "Costo flete (Bs/viaje)",
    "Sucursal",
    "Capacidad en KG",
    "Capacidad de carga útil en maples",
    "Capacidad de carga útil en Kg",
    "Sistema Camión"
]

# ── Callback para el worker de la cola ─────────────────────────────────
async def write_callback(item: QueueItem):
    """
    Ejecuta la escritura real en Google Sheets (vía Apps Script o gspread).
    """
    from modules.camiones.db.database import marcar_sincronizado
    
    if not sheets_client.enabled:
        logger.warning("Se omitió el envío a Google Sheets para el item #%s porque el cliente no está habilitado.", item.auditoria_id)
        return

    if item.action == "append":
        result = await sheets_client.append_row(item.valores)
        if result.get("success"):
            fila_real = result.get("data", {}).get("fila_insertada") or result.get("data")
            await marcar_sincronizado(fila_id=item.fila_id, nuevo_fila_id_real=fila_real)
        else:
            raise Exception(result.get("error", "Error desconocido al hacer append"))
    elif item.action == "update_row":
        result = await sheets_client.update_row(item.fila_id, item.valores)
        if result.get("success"):
            await marcar_sincronizado(fila_id=item.fila_id)
        else:
            raise Exception(result.get("error", "Error desconocido al hacer update"))

    await actualizar_estado_auditoria(auditoria_id=item.auditoria_id, estado="éxito")

# Cola asíncrona
update_queue = UpdateQueue(
    write_callback=write_callback,
    max_retries=settings.max_retries,
    retry_base_delay=settings.retry_base_delay,
    rate_limit_max=settings.rate_limit_max,
    rate_limit_window=settings.rate_limit_window,
)

router = APIRouter(tags=["Camiones"])

# ── Endpoints API ──────────────────────────────────────────────────────

@router.post("/api/push-to-sheets")
async def push_to_sheets():
    """
    Sube TODOS los datos locales (SQLite) al Google Sheet en background.
    """
    if not sheets_client.enabled:
        raise HTTPException(400, "Google Sheets no está configurado")
    from modules.camiones.lifecycle import _push_task, push_to_sheets_background
    if _push_task and not _push_task.done():
        return {"success": True, "message": "Ya hay una subida en progreso."}
    _push_task = asyncio.create_task(push_to_sheets_background())
    return {"success": True, "message": "Subida iniciada en segundo plano. Esperá unos minutos y revisá el Sheet."}

@router.get("/api/push-status")
async def push_status():
    """Estado de la subida actual."""
    from modules.camiones.lifecycle import _push_task
    if _push_task and not _push_task.done():
        return {"success": True, "running": True, "message": "Subida en progreso..."}
    return {"success": True, "running": False, "message": "Sin subida activa."}

@router.get("/api/health")
async def health():
    """Health check simple para monitoreo."""
    return {"status": "ok", "mode": "sheets" if sheets_client.enabled else "local"}

@router.get("/api/dashboard/stats")
async def dashboard_stats():
    """Estadísticas completas para el dashboard (KPIs, promedios, alertas)."""
    from modules.camiones.db.database import (
        obtener_todos_camiones, obtener_promedios_ruta,
        obtener_promedio_flete_por_sucursal
    )
    camiones = await obtener_todos_camiones()
    promedios_ruta = await obtener_promedios_ruta()
    promedios_sucursal = await obtener_promedio_flete_por_sucursal()

    total = len(camiones)
    en_servicio = sum(1 for c in camiones if (c.estado_servicio or "EN SERVICIO").upper() == "EN SERVICIO")
    fuera_servicio = sum(1 for c in camiones if (c.estado_servicio or "").upper() == "FUERA DE SERVICIO")
    consultar = sum(1 for c in camiones if (c.estado_servicio or "").upper() == "CONSULTAR")
    total_capacidad_maples = sum(c.capacidad_maples or 0 for c in camiones)
    total_capacidad_kg = sum(c.capacidad_util_kg or 0 for c in camiones)

    # Calcular flete total desde precios por ruta
    ruta_precios = {p["ruta"]: p["promedio"] for p in promedios_ruta if p["promedio"] > 0}
    total_flete = 0.0
    camiones_con_precio_ruta = 0
    for c in camiones:
        ruta = c.ruta or "local"
        precio = ruta_precios.get(ruta, 0)
        if precio > 0:
            total_flete += precio
            camiones_con_precio_ruta += 1

    # Camiones sin ruta asignada
    sin_ruta = sum(1 for c in camiones if not c.ruta or c.ruta in ("", "-"))
    con_ruta = total - sin_ruta

    # Promedio general de flete (promedio simple de precios por ruta)
    precios_activos = [p["promedio"] for p in promedios_ruta if p["promedio"] > 0]
    prom_general = round(sum(precios_activos) / len(precios_activos), 2) if precios_activos else 0

    # Ruta más rentable (mayor promedio)
    ruta_mas_rentable = max(promedios_ruta, key=lambda p: p["promedio"]) if promedios_ruta else None

    # Rutas madres más frecuentes
    from collections import Counter
    madres = Counter(c.ruta_madre or "(sin ruta)" for c in camiones)
    top_rutas = [{"ruta": r, "count": n} for r, n in madres.most_common(10)]

    # Por sucursal: proyección desde precios por ruta
    suc_map = {}
    for c in camiones:
        s = c.sucursal or "Sin Sucursal"
        if s not in suc_map:
            suc_map[s] = {"cantidad": 0, "total_flete": 0, "camiones_con_flete": 0}
        suc_map[s]["cantidad"] += 1
        ruta = c.ruta or "local"
        precio = ruta_precios.get(ruta, 0)
        if precio > 0:
            suc_map[s]["total_flete"] += precio
            suc_map[s]["camiones_con_flete"] += 1
    suc_resumen = [
        {"sucursal": s, "cantidad": v["cantidad"], "total_flete": round(v["total_flete"], 2),
         "promedio": round(v["total_flete"] / max(v.get("camiones_con_flete", 1), 1), 2),
         "camiones_con_flete": v.get("camiones_con_flete", 0)}
        for s, v in sorted(suc_map.items())
    ]

    # Alertas
    alertas = []
    if fuera_servicio > 0:
        alertas.append({"tipo": "warning", "mensaje": f"{fuera_servicio} camión(es) están FUERA DE SERVICIO"})
    if consultar > 0:
        alertas.append({"tipo": "info", "mensaje": f"{consultar} camión(es) están en estado CONSULTAR"})
    rutas_sin_precio = total - camiones_con_precio_ruta
    if rutas_sin_precio > 0:
        alertas.append({"tipo": "info", "mensaje": f"{rutas_sin_precio} camión(es) sin precio de ruta asignado"})
    if sin_ruta > 0:
        alertas.append({"tipo": "info", "mensaje": f"{sin_ruta} camión(es) sin ruta asignada"})

    return {
        "kpis": {
            "total_camiones": total,
            "en_servicio": en_servicio,
            "fuera_servicio": fuera_servicio,
            "consultar": consultar,
            "total_capacidad_maples": total_capacidad_maples,
            "total_capacidad_kg": total_capacidad_kg,
            "total_flete": round(total_flete, 2),
            "promedio_flete_general": prom_general,
            "camiones_con_flete": camiones_con_precio_ruta,
            "sin_ruta": sin_ruta,
            "con_ruta": con_ruta,
            "ruta_mas_rentable": {"ruta": ruta_mas_rentable["ruta"], "promedio": ruta_mas_rentable["promedio"]} if ruta_mas_rentable else None,
        },
        "resumen_sucursales": suc_resumen,
        "promedios_ruta": [
            {
                "ruta": p["ruta"],
                "cantidad_viajes": p["cantidad_viajes"],
                "total_pagado": p["total_pagado"],
                "promedio": p["promedio"],
            }
            for p in promedios_ruta
        ],
        "promedios_sucursal": [
            {
                "sucursal": s["sucursal"],
                "promedio_flete": s["promedio_flete"],
                "total_camiones": s["total_camiones"],
                "total_flete": s["total_flete"],
            }
            for s in promedios_sucursal
        ],
        "top_rutas_madre": top_rutas,
        "alertas": alertas,
    }

# ── Auth ───────────────────────────────────────────────────────
from pydantic import BaseModel

class LoginRequest(BaseModel):
    username: str
    password: str

@router.post("/api/login")
async def login(req: LoginRequest):
    if verify_credentials(req.username, req.password):
        token = create_session(req.username)
        session = verify_session(token)
        return {"success": True, "token": token, "username": req.username, "display_name": session["display_name"]}
    return {"success": False, "message": "Credenciales inválidas"}

@router.post("/api/logout")
async def logout(token: str = ""):
    destroy_session(token)
    return {"success": True}

@router.get("/api/check-auth")
async def check_auth(token: str = ""):
    session = verify_session(token)
    if session:
        return {"success": True, "username": session["username"], "display_name": session["display_name"]}
    return {"success": False}

@router.get("/api/status", response_model=SyncStatusResponse)
async def get_status():
    """
    Obtiene el estado de la conexión a Sheets y la cola de procesamiento.
    """
    pendientes = await obtener_pendientes_sincronizacion_count()
    total = await obtener_total_camiones_count()
    ultimo_cambio, ultimo_por, ultimo_email = await obtener_ultimo_cambio()
    por_sucursal = await obtener_camiones_por_sucursal()
    
    return SyncStatusResponse(
        modo="sheets" if sheets_client.enabled else "local",
        sheets_configuradas=sheets_client.enabled,
        pendientes_sincronizacion=pendientes,
        total_registros=total,
        ultimo_cambio=ultimo_cambio,
        ultimo_cambio_por=ultimo_por,
        ultimo_cambio_email=ultimo_email,
        camiones_por_sucursal=por_sucursal
    )

@router.get("/api/camiones", response_model=list[CamionResponse])
async def list_camiones():
    """
    Lista todos los camiones desde la caché de SQLite (rápido).
    """
    try:
        return await obtener_todos_camiones()
    except Exception as e:
        logger.error("Error al listar camiones: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/camiones/export/xlsx")
async def export_camiones_xlsx(
    placa: str = "",
    sucursal: str = "",
    tipo_combustible: str = "",
    sistema_camion: str = "",
    estado_servicio: str = ""
):
    """
    Exporta camiones a Excel (.xlsx) aplicando los mismos filtros del dashboard.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    todos = await obtener_todos_camiones()
    promedios_map = {p["ruta"]: p["promedio"] for p in await obtener_promedios_ruta()}
    query = placa.strip().lower()
    camiones = [
        c for c in todos
        if (not query or query in c.placa.lower())
        and (not sucursal or c.sucursal == sucursal)
        and (not tipo_combustible or c.tipo_combustible == tipo_combustible)
        and (not sistema_camion or c.sistema_camion == sistema_camion)
        and (not estado_servicio or (c.estado_servicio or "EN SERVICIO") == estado_servicio)
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Camiones"

    headers = ["Placa","Sucursal","Sistema","Servicio","Ruta Madre","Ruta","Flete (Bs)","Capacidad (Maples)"]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="Calibri", size=10)
    for i, c in enumerate(camiones, 2):
        ruta_key = c.ruta or "local"
        flete_val = promedios_map.get(ruta_key, c.costo_flete or 0)
        vals = [
            c.placa, c.sucursal, c.sistema_camion or "SIN INFORMACIÓN",
            c.estado_servicio or "EN SERVICIO", c.ruta_madre or "",
            c.ruta or "local", flete_val,
            c.capacidad_maples or 0,
        ]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = data_font
            cell.border = thin_border

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 14

    from fastapi.responses import StreamingResponse
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=camiones_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"}
    )

@router.post("/api/camiones", response_model=UpdateSheetResponse)
async def create_camion(request: CamionCreate):
    """
    Registra un nuevo camión.
    Calcula el Nº según la sucursal y la fila_id según el max local.
    Sincroniza en background si Sheets está habilitado.
    """
    logger.info("Registrando nuevo camión con placa: %s", request.placa)
    
    # Verificar placa duplicada
    duplicado = await obtener_camion_por_placa(request.placa)
    if duplicado:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"La placa '{request.placa}' ya se encuentra registrada."
        )
        
    try:
        # 1. Auto-calcular correlativo Nº por sucursal
        siguiente_nro = await obtener_siguiente_nro_sucursal(request.sucursal)
        
        # 2. Auto-calcular capacidad útil en Kg si no se especificó o para consistencia (factor 1.95)
        capacidad_util_calculada = round(request.capacidad_maples * 1.95, 2)
        
        # 3. Auto-calcular el fila_id
        max_fila = await obtener_max_fila_id()
        nueva_fila = max_fila + 1
        
        camion_dict = request.model_dump()
        camion_dict["nro"] = str(siguiente_nro)
        camion_dict["capacidad_util_kg"] = capacidad_util_calculada
        
        # Estado inicial
        estado_sinc = "pendiente_insercion" if sheets_client.enabled else "local"
        
        # Guardar localmente de inmediato
        await crear_camion_local(camion_dict, nueva_fila, estado_sinc)
        
        auditoria_id = None
        if sheets_client.enabled:
            valores_fila = [
                str(siguiente_nro),
                request.placa,
                request.estado_trabajo,
                request.tipo_combustible,
                str(request.costo_flete),
                request.sucursal,
                str(request.capacidad_kg),
                str(request.capacidad_maples),
                str(capacidad_util_calculada),
                request.sistema_camion,
                request.estado_servicio or "EN SERVICIO",
                request.ruta_madre or "",
                request.ruta or "",
            ]
            
            # Registrar auditoría
            auditoria_id = await crear_registro_auditoria(
                fila_id=nueva_fila,
                accion="crear",
                valores=json.dumps(camion_dict)
            )
            
            # Encolar
            item = QueueItem(
                auditoria_id=auditoria_id,
                action="append",
                fila_id=nueva_fila,
                valores=valores_fila
            )
            await update_queue.enqueue(item)
            
            asyncio.create_task(recalcular_promedios_ruta())
            return UpdateSheetResponse(
                success=True,
                message=f"Registro guardado localmente y encolado para Google Sheets (auditoría #{auditoria_id}).",
                auditoria_id=auditoria_id
            )
        else:
            # Recalcular promedios por ruta
            asyncio.create_task(recalcular_promedios_ruta())
            return UpdateSheetResponse(
                success=True,
                message="Registro guardado localmente de forma exitosa (Modo Local).",
                auditoria_id=None
            )
            
    except Exception as e:
        logger.error("Error al registrar camión: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/api/camiones/{fila_id}", response_model=UpdateSheetResponse)
async def update_camion(fila_id: int, request: CamionUpdate):
    """
    Modifica los detalles de un camión existente.
    Sincroniza en background si Sheets está habilitado.
    """
    logger.info("Modificando camión de la fila_id: %d", fila_id)
    
    camion_existente = await obtener_camion_por_fila(fila_id)
    if not camion_existente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Camión en la fila {fila_id} no encontrado."
        )
        
    try:
        camion_dict = request.model_dump(exclude_unset=True)
        
        # Si se modificaron los maples, recalcular capacidad útil
        if "capacidad_maples" in camion_dict:
            camion_dict["capacidad_util_kg"] = round(camion_dict["capacidad_maples"] * 1.95, 2)
            
        estado_sinc = "pendiente_actualizacion" if sheets_client.enabled else "local"
        
        # Guardar en base local
        camion_actualizado = await actualizar_camion_local(fila_id, camion_dict, estado_sinc)
        
        auditoria_id = None
        if sheets_client.enabled and camion_actualizado:
            valores_fila = [
                str(camion_actualizado.nro or ""),
                str(camion_actualizado.placa),
                str(camion_actualizado.estado_trabajo),
                str(camion_actualizado.tipo_combustible),
                str(camion_actualizado.costo_flete),
                str(camion_actualizado.sucursal),
                str(camion_actualizado.capacidad_kg),
                str(camion_actualizado.capacidad_maples),
                str(camion_actualizado.capacidad_util_kg),
                str(camion_actualizado.sistema_camion),
                str(camion_actualizado.estado_servicio or "EN SERVICIO"),
                str(camion_actualizado.ruta_madre or ""),
                str(camion_actualizado.ruta or ""),
            ]
            
            # Registrar auditoría
            auditoria_id = await crear_registro_auditoria(
                fila_id=fila_id,
                accion="editar",
                valores=json.dumps(camion_dict)
            )
            
            # Encolar
            item = QueueItem(
                auditoria_id=auditoria_id,
                action="update_row",
                fila_id=fila_id,
                valores=valores_fila
            )
            await update_queue.enqueue(item)
            
            asyncio.create_task(recalcular_promedios_ruta())
            return UpdateSheetResponse(
                success=True,
                message=f"Modificación guardada localmente y encolada para Google Sheets (auditoría #{auditoria_id}).",
                auditoria_id=auditoria_id
            )
        else:
            # Recalcular promedios por ruta
            asyncio.create_task(recalcular_promedios_ruta())
            return UpdateSheetResponse(
                success=True,
                message="Modificación guardada localmente de forma exitosa (Modo Local).",
                auditoria_id=None
            )
            
    except Exception as e:
        logger.error("Error al actualizar camión: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/api/camiones/{fila_id}", response_model=UpdateSheetResponse)
async def delete_camion(fila_id: int):
    """
    Elimina un camión de SQLite y lo encola para borrar de Sheets.
    """
    logger.info("Eliminando camión fila_id: %d", fila_id)
    
    camion = await obtener_camion_por_fila(fila_id)
    if not camion:
        raise HTTPException(status_code=404, detail="Camión no encontrado.")
    
    try:
        if sheets_client.enabled:
            result = await sheets_client.delete_row(fila_id)
            if not result.get("success"):
                logger.error("Delete en Sheets falló: %s", result.get("error"))
                return UpdateSheetResponse(
                    success=False,
                    message=f"No se pudo eliminar en Google Sheets: {result.get('error', 'error desconocido')}",
                    auditoria_id=None
                )
            # Ajustar fila_id local para los registros que estaban debajo
            from modules.camiones.db.database import CamionDb, async_session_factory
            async with async_session_factory() as session:
                stmt = select(CamionDb).where(CamionDb.fila_id > fila_id).order_by(CamionDb.fila_id)
                rows_to_shift = await session.execute(stmt)
                for row in rows_to_shift.scalars():
                    row.fila_id -= 1
                await session.commit()
        
        await eliminar_camion_local(fila_id)
        
        auditoria_id = None
        if sheets_client.enabled:
            vals = {"placa": camion.placa}
            if camion.modificado_por:
                vals["eliminado_por"] = camion.modificado_por
                vals["modificado_por_email"] = camion.modificado_por_email or ""
            auditoria_id = await crear_registro_auditoria(
                fila_id=fila_id, accion="eliminar",
                valores=json.dumps(vals)
            )
        
        asyncio.create_task(recalcular_promedios_ruta())
        return UpdateSheetResponse(
            success=True,
            message=f"Camión {camion.placa} eliminado correctamente.",
            auditoria_id=auditoria_id
        )
    except Exception as e:
        logger.error("Error al eliminar camión: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/sync")
async def force_sync():
    """
    Forzar sincronización desde Google Sheets usando UPSERT.
    Así los camiones agregados localmente no se pierden.
    """
    if not sheets_client.enabled:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Google Sheets no está configurado o está deshabilitado."
        )

    try:
        result = await sheets_client.read_all_rows()
        if not result.get("success"):
            return {"success": False, "message": result.get("error", "Error del Apps Script")}

        rows = result.get("data", [])
        if not rows:
            return {"success": False, "message": "No hay datos en el Sheet."}

        camiones = []
        for obj in rows:
            entry = {
                "fila_id": obj.get("fila_id", 0),
                "nro": str(obj.get("nro", "")),
                "placa": str(obj.get("placa", "")),
                "estado_trabajo": str(obj.get("estado_trabajo", "Fijo")),
                "tipo_combustible": str(obj.get("tipo_combustible", "GAS-GASOLINA")),
                "costo_flete": float(obj.get("costo_flete", 0)),
                "sucursal": str(obj.get("sucursal", "")),
                "capacidad_kg": int(obj.get("capacidad_kg", 0)),
                "capacidad_maples": int(obj.get("capacidad_maples", 0)),
                "capacidad_util_kg": float(obj.get("capacidad_util_kg", 0)),
            }
            raw_sistema = (obj.get("sistema_camion") or "").strip().upper()
            entry["sistema_camion"] = "SIN INFORMACIÓN" if raw_sistema in ("-", "—", "") else raw_sistema
            entry["estado_servicio"] = obj.get("estado_servicio") or "EN SERVICIO"
            entry["ruta_madre"] = obj.get("ruta_madre") or ""
            entry["ruta"] = obj.get("ruta") or "local"
            camiones.append(entry)

        # Preserve local fields over sheet values for existing records
        camiones_local = {c.placa: c for c in await obtener_todos_camiones()}
        for c in camiones:
            placa = c["placa"]
            if placa in camiones_local:
                local = camiones_local[placa]
                c["sistema_camion"] = getattr(local, "sistema_camion", "SIN INFORMACIÓN")
                c["estado_servicio"] = getattr(local, "estado_servicio", "EN SERVICIO")
                c["ruta_madre"] = getattr(local, "ruta_madre", "")
                c["ruta"] = getattr(local, "ruta", "local")

        await upsert_camiones_desde_sheets(camiones)
        asyncio.create_task(recalcular_promedios_ruta())
        return {"success": True, "message": f"Sincronizados/actualizados {len(camiones)} camiones."}
    except Exception as e:
        logger.error("Error en sincronización forzada: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/fletes", response_model=list[FletePromedioResponse])
async def list_fletes():
    """
    Obtiene el promedio de flete por sucursal.
    """
    try:
        data = await obtener_promedio_flete_por_sucursal()
        return [FletePromedioResponse(**d) for d in data]
    except Exception as e:
        logger.error("Error al obtener fletes: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/fletes/ruta", response_model=list[PromedioFleteRutaResponse])
async def list_fletes_ruta():
    """
    Obtiene el promedio de flete por ruta (calculado desde costo_flete de camiones).
    """
    try:
        data = await obtener_promedios_ruta()
        return [PromedioFleteRutaResponse(**d) for d in data]
    except Exception as e:
        logger.error("Error al obtener promedios por ruta: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/fletes/recalcular")
async def recalculate_fletes_ruta():
    """
    Recalcula todos los promedios de flete por ruta desde los camiones.
    """
    try:
        await recalcular_promedios_ruta()
        return {"success": True, "message": "Promedios por ruta recalculados."}
    except Exception as e:
        logger.error("Error al recalcular promedios: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/api/fletes/ruta/{ruta}")
async def update_flete_ruta(ruta: str, body: RutaPrecioUpdate):
    """Actualiza el precio de flete de una ruta manualmente."""
    try:
        await guardar_promedio_ruta(ruta, body.promedio)
        return {"success": True, "message": f"Precio de ruta '{ruta}' actualizado a Bs {body.promedio:.2f}"}
    except Exception as e:
        logger.error("Error al actualizar precio de ruta: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/fletes/ruta")
async def create_flete_ruta(body: RutaPrecioUpdate, ruta: str = ""):
    """Crea un nuevo precio de flete para una ruta (query param ?ruta=XXX)."""
    if not ruta:
        raise HTTPException(status_code=400, detail="Debe especificar ?ruta= en la URL")
    try:
        await guardar_promedio_ruta(ruta, body.promedio)
        return {"success": True, "message": f"Precio para ruta '{ruta}' creado: Bs {body.promedio:.2f}"}
    except Exception as e:
        logger.error("Error al crear precio de ruta: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/api/fletes/ruta/{ruta}")
async def delete_flete_ruta(ruta: str):
    """Elimina el precio de flete de una ruta."""
    try:
        await eliminar_promedio_ruta(ruta)
        return {"success": True, "message": f"Precio de ruta '{ruta}' eliminado."}
    except Exception as e:
        logger.error("Error al eliminar precio de ruta: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/bootstrap")
async def run_bootstrap():
    """
    Lee los 3 archivos HTML (LISTA CAMIONES 2024.html, DB.html, Hoja1.html),
    los mergea sin duplicados ni ruteos, y escribe en Google Sheets.
    """
    try:
        result = await bootstrap_sheets(force_write=True)
        return {
            "success": True,
            "total": len(result),
            "message": f"Bootstrap completado: {len(result)} camiones escritos en Sheets."
        }
    except Exception as e:
        logger.error("Error en bootstrap: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/auditoria", response_model=list[AuditEntry])
async def list_auditoria(limit: int = 20):
    """
    Obtiene los registros de auditoría recientes.
    Incluye quién hizo el cambio (extraído del JSON valores) y
    convierte la fecha a hora de Bolivia (UTC-4).
    """
    from datetime import timezone as tz, timedelta
    BOLIVIA_TZ = tz(timedelta(hours=-4))

    try:
        registros = await obtener_historial(limit=limit)

        response_list = []
        for r in registros:
            creado_bolivia = r.creado_en.astimezone(BOLIVIA_TZ)
            creado_str = creado_bolivia.strftime("%d/%m/%Y %H:%M:%S")

            # Extraer modificado_por y modificado_por_email del JSON valores
            mpor = None
            memail = None
            try:
                vals = json.loads(r.valores)
                if isinstance(vals, dict):
                    mpor = vals.get("modificado_por") or vals.get("eliminado_por") or None
                    memail = vals.get("modificado_por_email") or None
            except (json.JSONDecodeError, TypeError):
                pass

            response_list.append(
                AuditEntry(
                    id=r.id,
                    fila_id=r.fila_id,
                    accion=r.accion,
                    valores=r.valores,
                    estado=r.estado,
                    error=r.error,
                    creado_en=creado_str,
                    modificado_por=mpor,
                    modificado_por_email=memail
                )
            )
        return response_list
    except Exception as e:
        logger.error("Error al listar auditoría: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

# ── Tarifas de Flete ──────────────────────────────────────────────────

@router.get("/api/tarifas")
async def list_tarifas(clasificacion: str | None = None):
    """Lista todas las tarifas de flete. Opcional: ?clasificacion=EL ALTO"""
    try:
        return await obtener_todas_tarifas(clasificacion=clasificacion)
    except Exception as e:
        logger.error("Error al listar tarifas: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/api/tarifas/clasificaciones")
async def list_clasificaciones():
    """Lista las clasificaciones disponibles en las tarifas."""
    try:
        return await obtener_clasificaciones_tarifas()
    except Exception as e:
        logger.error("Error al listar clasificaciones: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/api/tarifas/load")
async def load_tarifas():
    """
    Carga las tarifas desde FLETE DE FLOTA DETALLE.xlsx.
    Reemplaza todas las tarifas existentes.
    """
    import openpyxl
    
    excel_path = Path(__file__).parent.parent.parent / "FLETE DE FLOTA DETALLE.xlsx"
    if not excel_path.exists():
        raise HTTPException(status_code=404, detail=f"No se encontro el archivo: {excel_path.name}")
    
    try:
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        ws = wb["Hoja1"]
        tarifas = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            clas = str(row[0]).strip() if row[0] else ""
            ruta = str(row[1]).strip() if row[1] else ""
            flete = float(row[2]) if row[2] is not None else 0.0
            if clas and ruta:
                tarifas.append({
                    "clasificacion": clas.upper(),
                    "tipo_ruta": ruta,
                    "flete_final": flete,
                })
        wb.close()
        
        if not tarifas:
            raise HTTPException(status_code=400, detail="No se encontraron datos validos en el Excel")
        
        count = await cargar_tarifas_desde_lista(tarifas)
        return {"success": True, "message": f"{count} tarifas cargadas exitosamente"}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error al cargar tarifas: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/api/tarifas/{tarifa_id}")
async def delete_tarifa(tarifa_id: int):
    """Elimina una tarifa por ID."""
    ok = await eliminar_tarifa(tarifa_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Tarifa no encontrada")
    return {"success": True, "message": f"Tarifa {tarifa_id} eliminada"}

# ── Rutas Madre/Hija (planilla de métodos de pago) ─────────

RUTAS_EXCEL_PATH = Path(__file__).parent.parent.parent / "PLANILLA DE MÉTODOS PARA PAGOS.xlsx"

@router.get("/api/rutas/seed")
async def seed_rutas():
    """Carga/actualiza la tabla de rutas desde el Excel."""
    if not RUTAS_EXCEL_PATH.exists():
        raise HTTPException(404, f"No se encontró: {RUTAS_EXCEL_PATH.name}")
    try:
        count = await seed_rutas_desde_excel(str(RUTAS_EXCEL_PATH))
        return {"success": True, "message": f"{count} rutas cargadas desde Excel"}
    except Exception as e:
        logger.error("Error al seed rutas: %s", e)
        raise HTTPException(500, detail=str(e))

@router.get("/api/rutas/madres")
async def list_rutas_madres():
    """Lista las rutas madre disponibles."""
    try:
        return await obtener_rutas_madres()
    except Exception as e:
        logger.error("Error al listar rutas madre: %s", e)
        raise HTTPException(500, detail=str(e))

@router.get("/api/rutas/hijas")
async def list_rutas_hijas(madre: str):
    """Lista las rutas hijas para una ruta madre. ?madre=MERCADO"""
    if not madre:
        raise HTTPException(400, detail="Parámetro 'madre' es requerido")
    try:
        return await obtener_rutas_hijas(madre)
    except Exception as e:
        logger.error("Error al listar rutas hijas: %s", e)
        raise HTTPException(500, detail=str(e))

