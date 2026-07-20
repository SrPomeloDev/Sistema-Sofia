"""
database.py — Inicialización de SQLite y operaciones CRUD para camiones y auditoría.
"""

import logging
from datetime import datetime, timezone
import json

from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column
from sqlalchemy import Integer, String, Text, DateTime, Float, select, func, delete

from modules.camiones.config import settings

logger = logging.getLogger(__name__)

# Engine y sesión asíncrona
engine = create_async_engine(settings.database_url, echo=False)
async_session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

class Base(DeclarativeBase):
    pass

class CamionDb(Base):
    """
    Tabla de camiones sincronizada localmente.
    """
    __tablename__ = "camiones"

    fila_id: Mapped[int] = mapped_column(Integer, primary_key=True) # Fila en Google Sheets (2, 3, 4...)
    nro: Mapped[str | None] = mapped_column(String(50), nullable=True) # Nº secuencial por sucursal
    placa: Mapped[str] = mapped_column(String(50), unique=True, index=True, nullable=False)
    estado_trabajo: Mapped[str] = mapped_column(String(50), default="Fijo")
    ruta: Mapped[str] = mapped_column(String(50), default="local")
    ruta_madre: Mapped[str] = mapped_column(String(100), default="")
    tipo_combustible: Mapped[str] = mapped_column(String(50), default="GAS-GASOLINA")
    costo_flete: Mapped[float] = mapped_column(Float, default=0.0)
    sucursal: Mapped[str] = mapped_column(String(100), nullable=False)
    capacidad_kg: Mapped[int] = mapped_column(Integer, default=0)
    capacidad_maples: Mapped[int] = mapped_column(Integer, default=0)
    capacidad_util_kg: Mapped[float] = mapped_column(Float, default=0.0)
    sistema_camion: Mapped[str] = mapped_column(String(50), default="SIN INFORMACIÓN")
    estado_servicio: Mapped[str] = mapped_column(String(20), default="EN SERVICIO")
    propietario: Mapped[str] = mapped_column(String(200), default="")
    modificado_por: Mapped[str] = mapped_column(String(100), default="")
    modificado_por_email: Mapped[str] = mapped_column(String(100), default="")
    
    @property
    def factor_0_75(self) -> float:
        return round((self.capacidad_maples or 0) * 0.75, 2)

    # Control de sincronización
    estado_sincronizacion: Mapped[str] = mapped_column(String(50), default="sincronizado") # 'sincronizado', 'pendiente_insercion', 'pendiente_actualizacion', 'local'
    error_sincronizacion: Mapped[str | None] = mapped_column(Text, nullable=True)
    actualizado_en: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        nullable=False,
        default=lambda: datetime.now(timezone.utc),
        onupdate=lambda: datetime.now(timezone.utc),
    )

class PromedioFleteRutaDb(Base):
    """
    Promedio de flete por ruta, calculado desde CamionDb.costo_flete.
    """
    __tablename__ = "promedios_ruta"

    ruta: Mapped[str] = mapped_column(String(100), primary_key=True)
    cantidad_viajes: Mapped[int] = mapped_column(Integer, default=0)
    total_pagado: Mapped[float] = mapped_column(Float, default=0.0)
    promedio: Mapped[float] = mapped_column(Float, default=0.0)

class RutaDb(Base):
    """
    Rutas Madre/Hija cargadas desde PLANILLA DE MÉTODOS PARA PAGOS.xlsx.
    Define la jerarquía: Ruta Madre (categoría) → Ruta Hija (asignación específica).
    """
    __tablename__ = "rutas"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    ruta_madre: Mapped[str] = mapped_column(String(100), nullable=False, index=True)
    ruta_hija: Mapped[str] = mapped_column(String(300), nullable=False)
    flete: Mapped[float] = mapped_column(Float, default=0.0)
    codigo_origen: Mapped[str | None] = mapped_column(String(50), nullable=True)
    observacion: Mapped[str | None] = mapped_column(Text, nullable=True)

class TarifaFleteDb(Base):
    """
    Tarifas de flete por clasificacion y tipo de ruta.
    Tabla de precios de referencia, cargada desde FLETE DE FLOTA DETALLE.xlsx.
    """
    __tablename__ = "tarifas_flete"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    clasificacion: Mapped[str] = mapped_column(String(100), nullable=False, index=True)
    tipo_ruta: Mapped[str] = mapped_column(String(300), nullable=False)
    flete_final: Mapped[float] = mapped_column(Float, default=0.0)

class Auditoria(Base):
    """
    Tabla de auditoría para operaciones con Google Sheets.
    """
    __tablename__ = "auditoria"

    id: Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    fila_id: Mapped[int | None] = mapped_column(Integer, nullable=True)
    accion: Mapped[str] = mapped_column(String(50), nullable=False) # "crear", "editar"
    valores: Mapped[str] = mapped_column(Text, nullable=False) # JSON string
    estado: Mapped[str] = mapped_column(String(20), nullable=False, default="pendiente") # "pendiente", "éxito", "fallido"
    error: Mapped[str | None] = mapped_column(Text, nullable=True)
    creado_en: Mapped[datetime] = mapped_column(
        DateTime(timezone=True),
        nullable=False,
        default=lambda: datetime.now(timezone.utc),
    )

# Inicialización
async def init_db():
    """
    Crea las tablas si no existen.
    """
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    logger.info("Base de datos SQLite inicializada correctamente.")

# CRUD de Camiones
async def obtener_todos_camiones() -> list[CamionDb]:
    async with async_session_factory() as session:
        # Ordenamos por sucursal y luego convertimos el nro a entero para ordenar numéricamente
        stmt = select(CamionDb).order_by(CamionDb.sucursal, CamionDb.fila_id)
        result = await session.execute(stmt)
        return list(result.scalars().all())

async def obtener_camion_por_placa(placa: str) -> CamionDb | None:
    async with async_session_factory() as session:
        stmt = select(CamionDb).where(CamionDb.placa == placa)
        result = await session.execute(stmt)
        return result.scalar_one_or_none()

async def obtener_camion_por_fila(fila_id: int) -> CamionDb | None:
    async with async_session_factory() as session:
        return await session.get(CamionDb, fila_id)

async def guardar_camiones_bulk(camiones_list: list[dict]):
    """
    Reemplaza TODO el contenido local con los datos del sheet.
    Deduplica por placa (conserva el primero).
    """
    async with async_session_factory() as session:
        await session.execute(delete(CamionDb))
        vistos = set()
        insertados = 0
        for item in camiones_list:
            placa = item["placa"]
            if placa in vistos:
                logger.warning("Placa duplicada ignorada en bulk insert: %s", placa)
                continue
            vistos.add(placa)
            session.add(CamionDb(
                fila_id=item["fila_id"],
                nro=item.get("nro"),
                placa=placa,
                estado_trabajo=item.get("estado_trabajo", "Fijo"),
                ruta=item.get("ruta", "local"),
                ruta_madre=item.get("ruta_madre", ""),
                tipo_combustible=item.get("tipo_combustible", "GAS-GASOLINA"),
                costo_flete=item.get("costo_flete", 0.0),
                sucursal=item["sucursal"],
                capacidad_kg=item.get("capacidad_kg", 0),
                capacidad_maples=item.get("capacidad_maples", 0),
                capacidad_util_kg=item.get("capacidad_util_kg", 0.0),
                sistema_camion=item.get("sistema_camion", "SIN INFORMACIÓN"),
                modificado_por=item.get("modificado_por", ""),
                modificado_por_email=item.get("modificado_por_email", ""),
                estado_servicio=item.get("estado_servicio", "EN SERVICIO"),
                propietario=item.get("propietario", ""),
                estado_sincronizacion="sincronizado"
            ))
            insertados += 1
        await session.commit()
        logger.info("Guardados %d camiones en SQLite de forma masiva (%d originales, %d duplicados ignorados)", insertados, len(camiones_list), len(camiones_list)-insertados)

async def upsert_camiones_desde_sheets(camiones_list: list[dict]):
    """
    Actualiza o inserta camiones respetando los que ya existen (por placa).
    Deduplica por placa (conserva el primero).
    Salta registros con placa vacía (garbage data).
    """
    async with async_session_factory() as session:
        # Obtener placas y fila_ids existentes
        existentes = await session.execute(select(CamionDb.placa, CamionDb.fila_id))
        placas_existentes = {}
        fila_ids_existentes = set()
        for row in existentes:
            placas_existentes[row[0]] = row[1]
            fila_ids_existentes.add(row[1])

        insertados = 0
        actualizados = 0
        vistos = set()
        for item in camiones_list:
            placa = item.get("placa", "").strip()
            if not placa:
                logger.warning("Placa vacía ignorada en upsert (fila_id=%s)", item.get("fila_id"))
                continue
            if placa in vistos:
                logger.warning("Placa duplicada ignorada en upsert: %s", placa)
                continue
            vistos.add(placa)
            if placa in placas_existentes:
                stmt = select(CamionDb).where(CamionDb.placa == placa)
                result = await session.execute(stmt)
                camion = result.scalar_one()
                # Solo actualizar fila_id si el nuevo valor no colisiona con otra fila
                nuevo_fila_id = item["fila_id"]
                if nuevo_fila_id != camion.fila_id and nuevo_fila_id in fila_ids_existentes:
                    logger.debug("fila_id %s ya existe para otra placa, conservando %s para placa %s",
                                 nuevo_fila_id, camion.fila_id, placa)
                else:
                    fila_ids_existentes.discard(camion.fila_id)
                    camion.fila_id = nuevo_fila_id
                    fila_ids_existentes.add(nuevo_fila_id)
                camion.nro = item.get("nro")
                camion.estado_trabajo = item.get("estado_trabajo", "Fijo")
                incoming_ruta = item.get("ruta", "")
                if incoming_ruta and incoming_ruta != "local":
                    camion.ruta = incoming_ruta
                incoming_madre = item.get("ruta_madre", "")
                if incoming_madre:
                    camion.ruta_madre = incoming_madre
                camion.tipo_combustible = item.get("tipo_combustible", "GAS-GASOLINA")
                camion.costo_flete = item.get("costo_flete", 0.0)
                camion.sucursal = item["sucursal"]
                camion.capacidad_kg = item.get("capacidad_kg", 0)
                camion.capacidad_maples = item.get("capacidad_maples", 0)
                camion.capacidad_util_kg = item.get("capacidad_util_kg", 0.0)
                camion.estado_sincronizacion = "sincronizado"
                actualizados += 1
            else:
                # Asignar fila_id que no colisione
                fila_id = item["fila_id"]
                while fila_id in fila_ids_existentes:
                    fila_id += 1
                fila_ids_existentes.add(fila_id)
                session.add(CamionDb(
                    fila_id=fila_id,
                    nro=item.get("nro"),
                    placa=placa,
                    estado_trabajo=item.get("estado_trabajo", "Fijo"),
                    ruta=item.get("ruta", "local"),
                    ruta_madre=item.get("ruta_madre", ""),
                    tipo_combustible=item.get("tipo_combustible", "GAS-GASOLINA"),
                    costo_flete=item.get("costo_flete", 0.0),
                    sucursal=item["sucursal"],
                    capacidad_kg=item.get("capacidad_kg", 0),
                    capacidad_maples=item.get("capacidad_maples", 0),
                    capacidad_util_kg=item.get("capacidad_util_kg", 0.0),
                    sistema_camion=item.get("sistema_camion", "SIN INFORMACIÓN"),
                    modificado_por=item.get("modificado_por", ""),
                    modificado_por_email=item.get("modificado_por_email", ""),
                    estado_servicio=item.get("estado_servicio", "EN SERVICIO"),
                    propietario=item.get("propietario", ""),
                    estado_sincronizacion="sincronizado"
                ))
                insertados += 1

        await session.commit()
        logger.info("Upsert desde Sheets: %d actualizados, %d insertados", actualizados, insertados)

        # ── Limpiar registros sincronizados que ya no existen en Sheets ───
        if vistos and len(vistos) > 5:
            stmt_delete = delete(CamionDb).where(
                CamionDb.estado_sincronizacion == "sincronizado",
                CamionDb.placa.notin_(vistos)
            )
            result = await session.execute(stmt_delete)
            if result.rowcount:
                logger.info("Eliminados %d camiones que ya no están en Sheets", result.rowcount)
                await session.commit()

async def crear_camion_local(camion_data: dict, fila_id: int, estado_sinc: str) -> CamionDb:
    async with async_session_factory() as session:
        camion = CamionDb(
            fila_id=fila_id,
            nro=camion_data.get("nro"),
            placa=camion_data["placa"],
            estado_trabajo=camion_data.get("estado_trabajo", "Fijo"),
            ruta=camion_data.get("ruta", "local"),
            ruta_madre=camion_data.get("ruta_madre", ""),
            tipo_combustible=camion_data.get("tipo_combustible", "GAS-GASOLINA"),
            costo_flete=camion_data.get("costo_flete", 0.0),
            sucursal=camion_data["sucursal"],
            capacidad_kg=camion_data.get("capacidad_kg", 0),
            capacidad_maples=camion_data.get("capacidad_maples", 0),
            capacidad_util_kg=camion_data.get("capacidad_util_kg", 0.0),
            sistema_camion=camion_data.get("sistema_camion", "SIN INFORMACIÓN"),
            estado_servicio=camion_data.get("estado_servicio", "EN SERVICIO"),
            propietario=camion_data.get("propietario", ""),
            modificado_por=camion_data.get("modificado_por", ""),
            modificado_por_email=camion_data.get("modificado_por_email", ""),
            estado_sincronizacion=estado_sinc
        )
        session.add(camion)
        await session.commit()
        await session.refresh(camion)
        return camion

async def actualizar_camion_local(fila_id: int, camion_data: dict, estado_sinc: str) -> CamionDb | None:
    async with async_session_factory() as session:
        camion = await session.get(CamionDb, fila_id)
        if not camion:
            return None
        
        for k, v in camion_data.items():
            if v is not None:
                setattr(camion, k, v)
        
        camion.estado_sincronizacion = estado_sinc
        await session.commit()
        await session.refresh(camion)
        return camion

async def marcar_sincronizado(fila_id: int, nuevo_fila_id_real: int | None = None):
    async with async_session_factory() as session:
        camion = await session.get(CamionDb, fila_id)
        if not camion:
            logger.warning("Camión con fila_id %s no encontrado para marcar como sincronizado", fila_id)
            return
        
        camion.estado_sincronizacion = "sincronizado"
        camion.error_sincronizacion = None
        
        if nuevo_fila_id_real is not None and nuevo_fila_id_real != fila_id:
            # Si se le asignó un nuevo ID real (por ejemplo, después de una inserción en Sheets)
            # Primero eliminamos el temporal e insertamos el real, o actualizamos la clave primaria
            # En SQLAlchemy actualizar la PK puede ser complejo; es más fácil recrearlo si cambia
            # Pero como calculamos fila_id = max_fila_id + 1 en el backend, nuevo_fila_id_real
            # suele ser idéntico al calculado. Si cambia, hacemos update del fila_id.
            # SQLite permite cambiar la PK directamente.
            camion.fila_id = nuevo_fila_id_real
            
        await session.commit()
        logger.debug("Camión fila %s marcado como sincronizado", fila_id)

async def marcar_error_sincronizacion(fila_id: int, error_msg: str):
    async with async_session_factory() as session:
        camion = await session.get(CamionDb, fila_id)
        if camion:
            camion.estado_sincronizacion = "error"
            camion.error_sincronizacion = error_msg
            await session.commit()
            logger.error("Camión fila %s marcado con error de sincronización: %s", fila_id, error_msg)

async def obtener_siguiente_nro_sucursal(sucursal: str) -> int:
    """Siguiente nro secuencial contando registros existentes de esa sucursal."""
    async with async_session_factory() as session:
        stmt = select(func.count()).select_from(CamionDb).where(CamionDb.sucursal == sucursal)
        result = await session.execute(stmt)
        return (result.scalar() or 0) + 1

async def obtener_max_fila_id() -> int:
    async with async_session_factory() as session:
        stmt = select(func.max(CamionDb.fila_id))
        result = await session.execute(stmt)
        val = result.scalar()
        return val if val is not None else 1 # Fila 1 es cabecera, así que retornamos 1 si no hay registros

async def obtener_pendientes_sincronizacion_count() -> int:
    async with async_session_factory() as session:
        stmt = select(func.count()).select_from(CamionDb).where(
            CamionDb.estado_sincronizacion.in_(["pendiente_insercion", "pendiente_actualizacion"])
        )
        result = await session.execute(stmt)
        return result.scalar() or 0

async def obtener_total_camiones_count() -> int:
    async with async_session_factory() as session:
        stmt = select(func.count()).select_from(CamionDb)
        result = await session.execute(stmt)
        return result.scalar() or 0

async def eliminar_camion_local(fila_id: int) -> bool:
    """Elimina un camión por fila_id. Retorna True si existía."""
    async with async_session_factory() as session:
        camion = await session.get(CamionDb, fila_id)
        if not camion:
            return False
        await session.delete(camion)
        await session.commit()
        logger.info("Camión fila_id %s eliminado", fila_id)
        return True

async def obtener_ultimo_cambio() -> tuple[str | None, str | None, str | None]:
    async with async_session_factory() as session:
        stmt = select(CamionDb.actualizado_en, CamionDb.modificado_por, CamionDb.modificado_por_email).order_by(CamionDb.actualizado_en.desc()).limit(1)
        result = await session.execute(stmt)
        row = result.first()
        if row:
            from datetime import timezone, timedelta
            utc_time = row[0]
            if utc_time:
                if utc_time.tzinfo is None:
                    utc_time = utc_time.replace(tzinfo=timezone.utc)
                bolivia_time = utc_time.astimezone(timezone(timedelta(hours=-4)))
                fecha = bolivia_time.strftime("%d/%m/%Y %H:%M:%S")
            else:
                fecha = None
            por = row[1] or "Sistema"
            email = row[2] or "sistema@sofia.bo"
            return fecha, por, email
        # Cuando no hay registros, devolver valores por defecto para evitar 'Sin registros'/None en la UI
        return None, "Sistema", "sistema@sofia.bo"

async def obtener_camiones_por_sucursal() -> dict[str, int]:
    async with async_session_factory() as session:
        stmt = select(CamionDb.sucursal, func.count()).group_by(CamionDb.sucursal).order_by(CamionDb.sucursal)
        result = await session.execute(stmt)
        return {row[0]: row[1] for row in result}

# ── Promedios por Ruta ─────────────────────────────────────────────

async def recalcular_promedios_ruta():
    """Recalcula la cantidad de viajes y el total pagado desde CamionDb,
    pero preserva el precio (promedio) si ya está registrado para esa ruta."""
    async with async_session_factory() as session:
        # Obtener todas las rutas que ya tienen precio manual o base
        existing = await session.execute(select(PromedioFleteRutaDb))
        existing_map = {p.ruta: p for p in existing.scalars()}

        stmt = select(
            CamionDb.ruta,
            func.count(),
            func.sum(CamionDb.costo_flete),
        ).where(CamionDb.costo_flete > 0).group_by(CamionDb.ruta)
        rows = await session.execute(stmt)
        
        rutas_actualizadas = set()
        for row in rows:
            ruta = row[0]
            cantidad = row[1]
            total = row[2] or 0.0
            if not ruta:
                continue
            rutas_actualizadas.add(ruta)
            
            if ruta in existing_map:
                p = existing_map[ruta]
                p.cantidad_viajes = cantidad
                p.total_pagado = round(total, 2)
                # Mantener el promedio manual existente
            else:
                promedio = round(total / cantidad, 2) if cantidad > 0 else 0.0
                session.add(PromedioFleteRutaDb(
                    ruta=ruta,
                    cantidad_viajes=cantidad,
                    total_pagado=round(total, 2),
                    promedio=promedio,
                ))
        
        # Para rutas que existen en el mapa pero no tienen viajes nuevos, poner viajes/pagado en 0
        for ruta, p in existing_map.items():
            if ruta not in rutas_actualizadas:
                p.cantidad_viajes = 0
                p.total_pagado = 0.0

        await session.commit()
        # Asegurar que se siembren todas las rutas de la planilla por si acaso
        await seed_precios_ruta_desde_planilla()
        logger.info("Promedios por ruta recalculados y sincronizados.")

async def obtener_promedios_ruta() -> list[dict]:
    async with async_session_factory() as session:
        stmt = select(PromedioFleteRutaDb).order_by(PromedioFleteRutaDb.ruta)
        result = await session.execute(stmt)
        return [
            {
                "ruta": p.ruta,
                "cantidad_viajes": p.cantidad_viajes,
                "total_pagado": p.total_pagado,
                "promedio": p.promedio,
            }
            for p in result.scalars()
        ]

async def obtener_promedio_por_ruta(ruta: str) -> dict | None:
    async with async_session_factory() as session:
        p = await session.get(PromedioFleteRutaDb, ruta)
        if not p:
            return None
        return {
            "ruta": p.ruta,
            "cantidad_viajes": p.cantidad_viajes,
            "total_pagado": p.total_pagado,
            "promedio": p.promedio,
        }

async def guardar_promedio_ruta(ruta: str, promedio: float):
    """Crea o actualiza el precio de una ruta manualmente."""
    async with async_session_factory() as session:
        existing = await session.get(PromedioFleteRutaDb, ruta)
        if existing:
            existing.promedio = round(promedio, 2)
            existing.cantidad_viajes = 0
            existing.total_pagado = 0.0
        else:
            session.add(PromedioFleteRutaDb(
                ruta=ruta,
                cantidad_viajes=0,
                total_pagado=0.0,
                promedio=round(promedio, 2),
            ))
        await session.commit()
        logger.info("Precio de ruta %s actualizado a Bs %.2f", ruta, promedio)

async def eliminar_promedio_ruta(ruta: str):
    """Elimina el precio de una ruta."""
    async with async_session_factory() as session:
        existing = await session.get(PromedioFleteRutaDb, ruta)
        if existing:
            await session.delete(existing)
            await session.commit()
            logger.info("Precio de ruta %s eliminado", ruta)

# ── Rutas Madre/Hija ──────────────────────────────────────────

async def seed_rutas_desde_excel(excel_path: str) -> int:
    """Carga/actualiza la tabla rutas desde el Excel de métodos de pago."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['Hoja1']
    async with async_session_factory() as session:
        await session.execute(delete(RutaDb))
        count = 0
        for row in range(2, ws.max_row + 1):
            madre = str(ws.cell(row, 5).value or '').strip()
            hija = str(ws.cell(row, 4).value or '').strip()
            flete_val = ws.cell(row, 9).value
            codigo = str(ws.cell(row, 3).value or '').strip()
            obs = str(ws.cell(row, 8).value or '').strip()
            if madre and hija:
                try:
                    flete = float(flete_val) if flete_val else 0.0
                except (ValueError, TypeError):
                    flete = 0.0
                session.add(RutaDb(
                    ruta_madre=madre,
                    ruta_hija=hija,
                    flete=flete,
                    codigo_origen=codigo or None,
                    observacion=obs or None,
                ))
                count += 1
        await session.commit()
        logger.info("Rutas sembradas desde Excel: %d registros (archivo: %s)", count, excel_path)
        return count

async def seed_precios_ruta_desde_planilla():
    """Siembra PromedioFleteRutaDb desde los fletes de RutaDb (PLANILLA).
    Solo agrega rutas que aún no tienen precio definido o siembra todas."""
    async with async_session_factory() as session:
        # Obtener todas las rutas hijas desde planilla
        stmt = select(RutaDb).order_by(RutaDb.ruta_hija)
        result = await session.execute(stmt)
        planilla_rutas = {(r.ruta_hija, r.flete) for r in result.scalars()}

        # Obtener rutas que ya tienen precio manual
        existing = await session.execute(select(PromedioFleteRutaDb.ruta))
        existing_rutas = {row[0] for row in existing}

        count = 0
        for ruta_hija, flete in planilla_rutas:
            if ruta_hija not in existing_rutas:
                session.add(PromedioFleteRutaDb(
                    ruta=ruta_hija,
                    cantidad_viajes=0,
                    total_pagado=0.0,
                    promedio=round(float(flete), 2),
                ))
                count += 1

        # También crear precios para rutas locales desde camiones sin precio
        stmt2 = select(CamionDb.ruta).distinct().where(
            CamionDb.ruta.isnot(None), CamionDb.ruta != "",
            CamionDb.ruta.notin_(list(existing_rutas) if existing_rutas else [""]),
        )
        result2 = await session.execute(stmt2)
        for row in result2:
            ruta = row[0]
            if ruta and ruta not in existing_rutas and ruta not in {r[0] for r in planilla_rutas}:
                session.add(PromedioFleteRutaDb(
                    ruta=ruta,
                    cantidad_viajes=0,
                    total_pagado=0.0,
                    promedio=0.0,
                ))
                count += 1

        await session.commit()
        if count > 0:
            logger.info("Precios de ruta sembrados desde planilla: %d nuevas rutas", count)

async def obtener_rutas_madres() -> list[str]:
    """Retorna lista ordenada de rutas madre únicas desde planilla + camiones.
    Filtra valores que no son rutas reales (RETIRADO, PLANIFICA)."""
    async with async_session_factory() as session:
        stmt = select(RutaDb.ruta_madre).distinct()
        result = await session.execute(stmt)
        madres = set(row[0] for row in result)
        stmt2 = select(CamionDb.ruta_madre).distinct()
        result2 = await session.execute(stmt2)
        for val in result2:
            if val[0] and val[0] not in ("RETIRADO", "PLANIFICA"):
                madres.add(val[0])
        return sorted(madres)

async def obtener_rutas_hijas(madre: str) -> list[dict]:
    """Retorna lista de rutas hijas para una ruta madre dada."""
    async with async_session_factory() as session:
        stmt = select(RutaDb).where(RutaDb.ruta_madre == madre).order_by(RutaDb.ruta_hija)
        result = await session.execute(stmt)
        return [
            {"id": r.id, "ruta_hija": r.ruta_hija, "flete": r.flete, "codigo_origen": r.codigo_origen}
            for r in result.scalars()
        ]

async def obtener_promedio_flete_por_sucursal() -> list[dict]:
    async with async_session_factory() as session:
        stmt = select(
            CamionDb.sucursal,
            func.avg(CamionDb.costo_flete),
            func.count(),
            func.sum(CamionDb.costo_flete)
        ).where(CamionDb.costo_flete > 0).group_by(CamionDb.sucursal).order_by(CamionDb.sucursal)
        result = await session.execute(stmt)
        return [
            {
                "sucursal": row[0],
                "promedio_flete": round(row[1] or 0, 2),
                "total_camiones": row[2],
                "total_flete": round(row[3] or 0, 2)
            }
            for row in result
        ]

# CRUD de Auditoría
async def crear_registro_auditoria(
    fila_id: int | None,
    accion: str,
    valores: str,
) -> int:
    async with async_session_factory() as session:
        registro = Auditoria(
            fila_id=fila_id,
            accion=accion,
            valores=valores,
            estado="pendiente",
        )
        session.add(registro)
        await session.commit()
        await session.refresh(registro)
        return registro.id

async def actualizar_estado_auditoria(
    auditoria_id: int,
    estado: str,
    error: str | None = None,
):
    async with async_session_factory() as session:
        registro = await session.get(Auditoria, auditoria_id)
        if registro:
            registro.estado = estado
            registro.error = error
            await session.commit()

async def obtener_historial(limit: int = 50) -> list[Auditoria]:
    async with async_session_factory() as session:
        stmt = (
            select(Auditoria)
            .order_by(Auditoria.creado_en.desc())
            .limit(limit)
        )
        result = await session.execute(stmt)
        return list(result.scalars().all())

# ── Tarifas de Flete ─────────────────────────────────────────────────

async def obtener_todas_tarifas(clasificacion: str | None = None) -> list[dict]:
    """Lista todas las tarifas, opcionalmente filtradas por clasificacion."""
    async with async_session_factory() as session:
        stmt = select(TarifaFleteDb).order_by(TarifaFleteDb.clasificacion, TarifaFleteDb.id)
        if clasificacion:
            stmt = stmt.where(TarifaFleteDb.clasificacion == clasificacion.strip().upper())
        result = await session.execute(stmt)
        return [
            {
                "id": t.id,
                "clasificacion": t.clasificacion,
                "tipo_ruta": t.tipo_ruta,
                "flete_final": t.flete_final,
            }
            for t in result.scalars()
        ]

async def cargar_tarifas_desde_lista(tarifas: list[dict]) -> int:
    """Reemplaza TODAS las tarifas con la lista dada. Retorna cantidad insertada."""
    async with async_session_factory() as session:
        await session.execute(delete(TarifaFleteDb))
        for item in tarifas:
            session.add(TarifaFleteDb(
                clasificacion=item["clasificacion"].strip().upper(),
                tipo_ruta=item["tipo_ruta"].strip(),
                flete_final=float(item["flete_final"]),
            ))
        await session.commit()
        return len(tarifas)

async def obtener_clasificaciones_tarifas() -> list[str]:
    """Lista las clasificaciones distintas en la tabla."""
    async with async_session_factory() as session:
        stmt = select(TarifaFleteDb.clasificacion).distinct().order_by(TarifaFleteDb.clasificacion)
        result = await session.execute(stmt)
        return [row[0] for row in result]

async def eliminar_tarifa(tarifa_id: int) -> bool:
    """Elimina una tarifa por ID. Retorna True si existia."""
    async with async_session_factory() as session:
        t = await session.get(TarifaFleteDb, tarifa_id)
        if not t:
            return False
        await session.delete(t)
        await session.commit()
        return True
