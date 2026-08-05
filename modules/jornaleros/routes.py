"""
routes.py — Endpoints REST del módulo Jornaleros.

Todos los endpoints mutantes requieren `?token=` válido (misma auth que
camiones/rutas). Las lecturas son públicas.
"""

import asyncio
import logging
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse

from modules.camiones.auth import verify_session
from modules.camiones.config import settings
from modules.jornaleros.models import (
    JornaleroCreate, JornaleroUpdate, JornaleroResponse,
    JornaleroListResponse, OperationResponse,
)
from modules.jornaleros.db.database import (
    get_all_jornaleros, get_jornalero_by_id, create_jornalero,
    update_jornalero, delete_jornalero, seed_desde_excel, limpiar_todo,
    obtener_stats, upsert_from_sheet_rows, obtener_todos, marcar_sincronizado,
)
from modules.jornaleros.sheets import jornaleros_sheets_client, HEADERS_LIST
from modules.jornaleros.services.queue import UpdateQueue, QueueItem

logger = logging.getLogger(__name__)
router = APIRouter(tags=["Jornaleros"])

_push_task: asyncio.Task | None = None
_push_result: dict | None = None


# ── Cola de escrituras a Sheets (misma lógica que camiones) ────────────
async def write_callback(item: QueueItem):
    """
    Ejecuta la escritura real en Google Sheets:
    - append: agrega una fila nueva al final.
    - update_row: busca el ID en la col A y actualiza ESA MISMA fila
      (si el ID ya no está en la hoja, la apenda).
    """
    from modules.jornaleros.db.database import marcar_sincronizado

    if not jornaleros_sheets_client.enabled:
        raise Exception("Cliente de Google Sheets no habilitado")

    if item.action == "append":
        result = await jornaleros_sheets_client.append_row(item.valores)
    elif item.action == "update_row":
        fila = await jornaleros_sheets_client._buscar_fila_por_id(item.jornalero_id)
        if fila is not None:
            result = await jornaleros_sheets_client.update_row(fila, item.valores)
        else:
            result = await jornaleros_sheets_client.append_row(item.valores)
    else:
        raise Exception(f"Acción desconocida: {item.action}")

    if result.get("success"):
        await marcar_sincronizado(item.jornalero_id)
    else:
        raise Exception(result.get("error", "Error desconocido al escribir en Sheets"))


update_queue = UpdateQueue(
    write_callback=write_callback,
    max_retries=settings.max_retries,
    retry_base_delay=settings.retry_base_delay,
    rate_limit_max=settings.rate_limit_max,
    rate_limit_window=settings.rate_limit_window,
)


def _require_auth(token: str):
    session = verify_session(token)
    if not session:
        raise HTTPException(status_code=401, detail="Sesión inválida o expirada")
    return session


def _a_response(row) -> JornaleroResponse:
    return JornaleroResponse(
        id=row.id,
        fecha_inicial=row.fecha_inicial,
        fecha_final=row.fecha_final,
        tipo_trabajador=row.tipo_trabajador,
        cd=row.cd,
        unidad=row.unidad,
        area=row.area,
        cantidad_jornaleros=row.cantidad_jornaleros,
        horas_trabajadas=row.horas_trabajadas,
        dias_trabajados_totales=row.dias_trabajados_totales,
        dias_trabajados_laborales=row.dias_trabajados_laborales,
        llenado_por=row.llenado_por,
        fecha_creacion=row.fecha_creacion,
        estado_sincronizacion=row.estado_sincronizacion,
        error_sincronizacion=row.error_sincronizacion,
        tarifa_diaria=row.tarifa_diaria,
        observaciones=row.observaciones,
        costo_total=row.costo_total,
    )


def _fecha_param(val: str | None) -> datetime | None:
    if not val:
        return None
    try:
        return datetime.fromisoformat(val.replace("Z", ""))
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Fecha inválida: {val}. Usá formato ISO (YYYY-MM-DD o YYYY-MM-DDTHH:MM:SS)")


@router.get("/api/jornaleros", response_model=JornaleroListResponse)
async def listar_jornaleros(
    offset: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    cd: str | None = None,
    area: str | None = None,
    tipo_trabajador: str | None = None,
    llenado_por: str | None = None,
    fecha_inicial__gte: str | None = None,
    fecha_inicial__lte: str | None = None,
    fecha_final__gte: str | None = None,
    fecha_final__lte: str | None = None,
):
    try:
        filas, total = await get_all_jornaleros(
            offset=offset, limit=limit, cd=cd, area=area,
            tipo_trabajador=tipo_trabajador, llenado_por=llenado_por,
            fecha_inicial_gte=_fecha_param(fecha_inicial__gte),
            fecha_inicial_lte=_fecha_param(fecha_inicial__lte),
            fecha_final_gte=_fecha_param(fecha_final__gte),
            fecha_final_lte=_fecha_param(fecha_final__lte),
        )
        return JornaleroListResponse(
            success=True,
            data=[_a_response(r) for r in filas],
            total=total, offset=offset, limit=limit,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error al listar jornaleros: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/jornaleros", response_model=OperationResponse)
async def crear_jornalero(body: JornaleroCreate, token: str = Query("")):
    _require_auth(token)
    try:
        row = await create_jornalero(body.model_dump())
        mensaje = "Registro de jornalero creado"
        if jornaleros_sheets_client.enabled:
            await update_queue.enqueue(QueueItem(
                jornalero_id=row.id,
                action="append",
                valores=_fila_a_lista(row),
            ))
            mensaje += " y enviado a Google Sheets (nueva fila)"
        else:
            mensaje += " (modo local, sin Sheets)"
        return OperationResponse(success=True, message=mensaje, data=_a_response(row))
    except Exception as e:
        logger.error("Error al crear jornalero: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/api/jornaleros/{jornalero_id}", response_model=OperationResponse)
async def actualizar_jornalero(jornalero_id: str, body: JornaleroUpdate, token: str = Query("")):
    _require_auth(token)
    try:
        datos = body.model_dump(exclude_unset=True)
        if not datos:
            raise HTTPException(status_code=400, detail="No hay campos para actualizar")
        row = await update_jornalero(jornalero_id, datos)
        if not row:
            raise HTTPException(status_code=404, detail="Registro de jornalero no encontrado")
        mensaje = "Registro actualizado"
        if jornaleros_sheets_client.enabled:
            await update_queue.enqueue(QueueItem(
                jornalero_id=row.id,
                action="update_row",
                valores=_fila_a_lista(row),
            ))
            mensaje += " y actualizada su misma fila en Google Sheets"
        else:
            mensaje += " (modo local, sin Sheets)"
        return OperationResponse(success=True, message=mensaje, data=_a_response(row))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error al actualizar jornalero %s: %s", jornalero_id, e)
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/api/jornaleros/{jornalero_id}", response_model=OperationResponse)
async def eliminar_jornalero(jornalero_id: str, token: str = Query("")):
    _require_auth(token)
    try:
        # 1. Borrar local primero (siempre, aunque sheets falle) — lógica de camiones
        ok = await delete_jornalero(jornalero_id)
        if not ok:
            raise HTTPException(status_code=404, detail="Registro de jornalero no encontrado")
        # 2. Borrar en Sheets por ID (best effort)
        if jornaleros_sheets_client.enabled:
            try:
                result = await jornaleros_sheets_client.delete_by_id(jornalero_id)
                if not result.get("success") and "no encontrado" not in str(result.get("error", "")).lower():
                    logger.warning("Delete en Sheets falló (local ya eliminado): %s", result.get("error"))
            except Exception as e:
                logger.warning("Delete en Sheets lanzó excepción (local ya eliminado): %s", e)
        return OperationResponse(success=True, message="Registro eliminado")
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error al eliminar jornalero %s: %s", jornalero_id, e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/jornaleros/sync")
async def sync_desde_sheets(token: str = Query("")):
    """Pull: trae todos los registros de Google Sheets y los upserta en la BD."""
    _require_auth(token)
    if not jornaleros_sheets_client.enabled:
        raise HTTPException(status_code=400, detail="Google Sheets no está configurado para jornaleros")
    try:
        await jornaleros_sheets_client.initialize()
        result = await jornaleros_sheets_client.read_all_rows()
        if not result.get("success"):
            raise HTTPException(status_code=502, detail=result.get("error", "Error al leer de Sheets"))
        datos = result.get("data") or []
        if datos and _es_fila_encabezados(datos[0]):
            resumen = await upsert_from_sheet_rows(datos)
        else:
            resumen = {"creados": 0, "actualizados": 0, "omitidos": len(datos)}
        return {"success": True, "message": "Sincronización completada", "detalles": resumen}
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error en sync de jornaleros: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


def _es_fila_encabezados(fila: list) -> bool:
    return any(str(c).strip().upper() == "ID" for c in fila) and any(str(c).strip().upper() == "CD" for c in fila)


def _fila_a_lista(row) -> list:
    """Convierte un registro de BD a la lista de 16 valores del sheet."""
    def _fmt_dt(val: datetime) -> str:
        return val.strftime("%Y-%m-%d") if val else ""
    return [
        row.id,
        _fmt_dt(row.fecha_inicial),
        _fmt_dt(row.fecha_final),
        row.tipo_trabajador or "JORNALERO",
        row.cd,
        row.unidad or "",
        row.area or "",
        row.cantidad_jornaleros or 0,
        row.horas_trabajadas or 0,
        row.dias_trabajados_totales or 0,
        row.dias_trabajados_laborales or 0,
        row.llenado_por or "",
        _fmt_dt(row.fecha_creacion),
        row.tarifa_diaria if row.tarifa_diaria is not None else "",
        row.observaciones or "",
    ]


async def _push_task_body():
    global _push_result
    try:
        # Reescritura completa en UN solo request (misma lógica que camiones):
        # headers + todas las filas → la hoja queda idéntica a la BD local.
        registros = await obtener_todos()
        valores = [_fila_a_lista(r) for r in registros]
        result = await jornaleros_sheets_client.set_all_rows(HEADERS_LIST, valores)
        if not result.get("success"):
            _push_result = {"total": len(registros), "sincronizados": 0, "fallidos": len(registros),
                            "error": result.get("error", "Error al escribir en la hoja")}
            return _push_result
        for r in registros:
            await marcar_sincronizado(r.id)
        _push_result = {"total": len(registros), "sincronizados": len(registros), "fallidos": 0}
        return _push_result
    except Exception as e:
        logger.error("Error en push de jornaleros: %s", e)
        _push_result = {"total": 0, "sincronizados": 0, "fallidos": 0, "error": str(e)}
        return _push_result


@router.post("/api/jornaleros/push-to-sheets")
async def push_a_sheets(token: str = Query("")):
    """Push: reescribe TODOS los registros locales en Google Sheets en segundo plano."""
    _require_auth(token)
    global _push_task, _push_result
    await jornaleros_sheets_client.initialize()
    if not jornaleros_sheets_client.enabled:
        raise HTTPException(status_code=400, detail="Google Sheets no está configurado para jornaleros (APPS_SCRIPT_URL/APPS_SCRIPT_TOKEN o credenciales OAuth)")
    if _push_task and not _push_task.done():
        return {"success": True, "running": True, "message": "Ya hay un push en progreso."}
    _push_result = None
    _push_task = asyncio.create_task(_push_task_body())
    return {"success": True, "running": True, "message": "Push iniciado en segundo plano."}


@router.get("/api/jornaleros/push-status")
async def push_status():
    global _push_task, _push_result
    if _push_task and not _push_task.done():
        return {"success": True, "running": True, "message": "Push en progreso..."}
    if _push_task and _push_task.done() and _push_task.exception():
        return {"success": False, "running": False, "message": f"Push falló: {_push_task.exception()}"}
    return {
        "success": True,
        "running": False,
        "message": "Push completado" if _push_result else "Sin push activo.",
        "result": _push_result,
    }


@router.get("/api/jornaleros/stats")
async def stats_jornaleros():
    try:
        return await obtener_stats()
    except Exception as e:
        logger.error("Error en stats de jornaleros: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/jornaleros/export/xlsx")
async def export_jornaleros_xlsx(
    token: str = Query(""),
    cd: str | None = None,
    area: str | None = None,
    fecha_inicial__gte: str | None = None,
    fecha_final__lte: str | None = None,
):
    """Exporta jornaleros filtrados a Excel (.xlsx). Requiere token."""
    _require_auth(token)
    try:
        filas, _ = await get_all_jornaleros(
            offset=0, limit=500, cd=cd, area=area,
            fecha_inicial_gte=_fecha_param(fecha_inicial__gte),
            fecha_final_lte=_fecha_param(fecha_final__lte),
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error en export jornaleros: %s", e)
        raise HTTPException(status_code=500, detail=str(e))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Jornaleros"

    headers = ["ID", "Fecha Inicial", "Fecha Final", "Tipo Trabajador", "CD", "Unidad",
               "Área", "Cant. Jornaleros", "Horas Trabajadas", "Días Totales",
               "Días Laborales", "Llenado por", "Fecha Creación", "Tarifa Diaria (Bs)",
               "Costo Total (Bs)", "Observaciones", "Estado Sync"]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, r in enumerate(filas, 2):
        valores = [
            r.id, r.fecha_inicial.strftime("%Y-%m-%d"), r.fecha_final.strftime("%Y-%m-%d"),
            r.tipo_trabajador, r.cd, r.unidad, r.area,
            r.cantidad_jornaleros, r.horas_trabajadas, r.dias_trabajados_totales,
            r.dias_trabajados_laborales, r.llenado_por,
            r.fecha_creacion.strftime("%Y-%m-%d %H:%M") if r.fecha_creacion else "",
            r.tarifa_diaria if r.tarifa_diaria is not None else "",
            r.costo_total if r.costo_total is not None else "",
            r.observaciones or "",
            r.estado_sincronizacion,
        ]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    widths = [12, 14, 14, 16, 14, 26, 12, 14, 14, 12, 12, 14, 18, 14, 14, 24, 14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=jornaleros.xlsx"},
    )


@router.post("/api/jornaleros/seed")
async def seed_jornaleros(token: str = Query(""), confirm: bool = Query(False)):
    """Importa la hoja Horas_Jornaleros de DB_PROD_GDN.xlsx a la BD."""
    _require_auth(token)
    if not confirm:
        raise HTTPException(status_code=400, detail="Debe enviar ?confirm=true para ejecutar el seed")
    try:
        resultado = await seed_desde_excel()
        return {"success": True, "message": f"Seed completado: {resultado['creados']} creados, {resultado['actualizados']} actualizados, {resultado['omitidos']} omitidos"}
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except KeyError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("Error en seed de jornaleros: %s", e)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/api/jornaleros/{jornalero_id}", response_model=OperationResponse)
async def obtener_jornalero(jornalero_id: str):
    try:
        row = await get_jornalero_by_id(jornalero_id)
        if not row:
            raise HTTPException(status_code=404, detail="Registro de jornalero no encontrado")
        return OperationResponse(success=True, message="OK", data=_a_response(row))
    except HTTPException:
        raise
    except Exception as e:
        logger.error("Error al obtener jornalero %s: %s", jornalero_id, e)
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/api/jornaleros/reseed")
async def reseed_jornaleros(token: str = Query(""), confirm: bool = Query(False)):
    """Limpia la tabla y re-importa desde el Excel."""
    _require_auth(token)
    if not confirm:
        raise HTTPException(status_code=400, detail="Debe enviar ?confirm=true para ejecutar el re-seed")
    try:
        await limpiar_todo()
        resultado = await seed_desde_excel()
        return {"success": True, "message": f"Re-seed completado: {resultado['creados']} creados, {resultado['actualizados']} actualizados"}
    except FileNotFoundError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except KeyError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.error("Error en reseed de jornaleros: %s", e)
        raise HTTPException(status_code=500, detail=str(e))
