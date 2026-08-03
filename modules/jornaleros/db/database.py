"""
db/database.py — Modelo SQLAlchemy y capa de acceso a datos del módulo Jornaleros.

La tabla `jornaleros` mapea exactamente la hoja `Horas_Jornaleros` del Excel
`DB_PROD_GDN.xlsx`. Cada registro usa el ID de la hoja como clave primaria
(cadena hex de 8 caracteres, ej: '3969a81d').
"""

import logging
import os
import uuid
from datetime import datetime, date

from sqlalchemy import String, Float, DateTime, Index, select, func, delete
from sqlalchemy.orm import DeclarativeBase, Mapped, mapped_column
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine, async_sessionmaker

from modules.camiones.config import settings

logger = logging.getLogger(__name__)

DATABASE_URL = os.getenv("DATABASE_URL") or settings.database_url or "sqlite+aiosqlite:///./auditoria.db"

engine = create_async_engine(
    DATABASE_URL, echo=False,
    connect_args={"check_same_thread": False}
)
async_session_factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)


class Base(DeclarativeBase):
    pass


class JornaleroDb(Base):
    __tablename__ = "jornaleros"

    id: Mapped[str] = mapped_column(String(36), primary_key=True)
    fecha_inicial: Mapped[datetime] = mapped_column(DateTime, nullable=False)
    fecha_final: Mapped[datetime] = mapped_column(DateTime, nullable=False)
    tipo_trabajador: Mapped[str] = mapped_column(String(20), default="JORNALERO", nullable=False)
    cd: Mapped[str] = mapped_column(String(50), nullable=False)
    unidad: Mapped[str] = mapped_column(String(100), default="", nullable=False)
    area: Mapped[str] = mapped_column(String(20), default="DESPACHO", nullable=False)
    cantidad_jornaleros: Mapped[float] = mapped_column(Float, default=0.0)
    horas_trabajadas: Mapped[float] = mapped_column(Float, default=0.0)
    dias_trabajados_totales: Mapped[float] = mapped_column(Float, default=0.0)
    dias_trabajados_laborales: Mapped[float] = mapped_column(Float, default=0.0)
    llenado_por: Mapped[str] = mapped_column(String(100), default="", nullable=False)
    fecha_creacion: Mapped[datetime] = mapped_column(DateTime, default=datetime.now)
    estado_sincronizacion: Mapped[str] = mapped_column(String(20), default="pendiente", nullable=False)
    error_sincronizacion: Mapped[str | None] = mapped_column(String(255), nullable=True)

    __table_args__ = (
        Index("ix_jornaleros_cd_area_fecha", "cd", "area", "fecha_inicial"),
    )


async def init_db():
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    logger.info("Tabla jornaleros creada/verificada")


# ── Helpers ────────────────────────────────────────────────────────────

def _nuevo_id() -> str:
    """ID hex de 8 caracteres, consistente con la hoja Horas_Jornaleros."""
    return uuid.uuid4().hex[:8]


def _parse_float_safe(val) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _parse_datetime_safe(val) -> datetime | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    try:
        return datetime.fromisoformat(str(val).strip().replace("Z", ""))
    except (ValueError, TypeError):
        return None


def _to_dict(row: JornaleroDb) -> dict:
    return {
        "id": row.id,
        "fecha_inicial": row.fecha_inicial,
        "fecha_final": row.fecha_final,
        "tipo_trabajador": row.tipo_trabajador,
        "cd": row.cd,
        "unidad": row.unidad,
        "area": row.area,
        "cantidad_jornaleros": row.cantidad_jornaleros,
        "horas_trabajadas": row.horas_trabajadas,
        "dias_trabajados_totales": row.dias_trabajados_totales,
        "dias_trabajados_laborales": row.dias_trabajados_laborales,
        "llenado_por": row.llenado_por,
        "fecha_creacion": row.fecha_creacion,
        "estado_sincronizacion": row.estado_sincronizacion,
        "error_sincronizacion": row.error_sincronizacion,
    }


def _resolve_seed_path() -> str:
    """Busca DB_PROD_GDN.xlsx en ubicaciones probables (env, raíz, Downloads)."""
    env_path = os.getenv("JORNALEROS_SEED_PATH", "")
    candidates = [
        env_path,
        os.path.join(os.getcwd(), "DB_PROD_GDN.xlsx"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", "..", "DB_PROD_GDN.xlsx"),
        os.path.join(os.path.expanduser("~"), "Downloads", "DB_PROD_GDN.xlsx"),
    ]
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[1] or "DB_PROD_GDN.xlsx"


# ── DAO: consultas ─────────────────────────────────────────────────────

async def get_all_jornaleros(
    offset: int = 0,
    limit: int = 100,
    cd: str | None = None,
    area: str | None = None,
    tipo_trabajador: str | None = None,
    llenado_por: str | None = None,
    fecha_inicial_gte: datetime | None = None,
    fecha_inicial_lte: datetime | None = None,
    fecha_final_gte: datetime | None = None,
    fecha_final_lte: datetime | None = None,
) -> tuple[list[JornaleroDb], int]:
    async with async_session_factory() as session:
        stmt = select(JornaleroDb)
        count_stmt = select(func.count(JornaleroDb.id))
        if cd:
            stmt = stmt.where(func.upper(JornaleroDb.cd) == cd.strip().upper())
            count_stmt = count_stmt.where(func.upper(JornaleroDb.cd) == cd.strip().upper())
        if area:
            stmt = stmt.where(func.upper(JornaleroDb.area) == area.strip().upper())
            count_stmt = count_stmt.where(func.upper(JornaleroDb.area) == area.strip().upper())
        if tipo_trabajador:
            stmt = stmt.where(func.upper(JornaleroDb.tipo_trabajador) == tipo_trabajador.strip().upper())
            count_stmt = count_stmt.where(func.upper(JornaleroDb.tipo_trabajador) == tipo_trabajador.strip().upper())
        if llenado_por:
            stmt = stmt.where(func.upper(JornaleroDb.llenado_por).like(f"%{llenado_por.strip().upper()}%"))
            count_stmt = count_stmt.where(func.upper(JornaleroDb.llenado_por).like(f"%{llenado_por.strip().upper()}%"))
        if fecha_inicial_gte:
            stmt = stmt.where(JornaleroDb.fecha_inicial >= fecha_inicial_gte)
            count_stmt = count_stmt.where(JornaleroDb.fecha_inicial >= fecha_inicial_gte)
        if fecha_inicial_lte:
            stmt = stmt.where(JornaleroDb.fecha_inicial <= fecha_inicial_lte)
            count_stmt = count_stmt.where(JornaleroDb.fecha_inicial <= fecha_inicial_lte)
        if fecha_final_gte:
            stmt = stmt.where(JornaleroDb.fecha_final >= fecha_final_gte)
            count_stmt = count_stmt.where(JornaleroDb.fecha_final >= fecha_final_gte)
        if fecha_final_lte:
            stmt = stmt.where(JornaleroDb.fecha_final <= fecha_final_lte)
            count_stmt = count_stmt.where(JornaleroDb.fecha_final <= fecha_final_lte)

        total = (await session.execute(count_stmt)).scalar() or 0
        stmt = stmt.order_by(JornaleroDb.fecha_inicial.desc(), JornaleroDb.cd).offset(offset).limit(limit)
        result = await session.execute(stmt)
        return list(result.scalars().all()), total


async def get_jornalero_by_id(jornalero_id: str) -> JornaleroDb | None:
    async with async_session_factory() as session:
        return await session.get(JornaleroDb, jornalero_id)


async def create_jornalero(data: dict) -> JornaleroDb:
    async with async_session_factory() as session:
        row = JornaleroDb(
            id=_nuevo_id(),
            fecha_inicial=data["fecha_inicial"],
            fecha_final=data["fecha_final"],
            tipo_trabajador=data.get("tipo_trabajador", "JORNALERO"),
            cd=data["cd"],
            unidad=data.get("unidad", ""),
            area=data.get("area", "DESPACHO"),
            cantidad_jornaleros=data.get("cantidad_jornaleros", 0.0),
            horas_trabajadas=data.get("horas_trabajadas", 0.0),
            dias_trabajados_totales=data.get("dias_trabajados_totales", 0.0),
            dias_trabajados_laborales=data.get("dias_trabajados_laborales", 0.0),
            llenado_por=data.get("llenado_por", ""),
            estado_sincronizacion="pendiente",
        )
        session.add(row)
        await session.commit()
        await session.refresh(row)
        return row


async def update_jornalero(jornalero_id: str, data: dict) -> JornaleroDb | None:
    async with async_session_factory() as session:
        row = await session.get(JornaleroDb, jornalero_id)
        if not row:
            return None
        for key, val in data.items():
            if key in ("id", "fecha_creacion"):
                continue
            setattr(row, key, val)
        row.estado_sincronizacion = "pendiente"
        row.error_sincronizacion = None
        await session.commit()
        await session.refresh(row)
        return row


async def delete_jornalero(jornalero_id: str) -> bool:
    async with async_session_factory() as session:
        row = await session.get(JornaleroDb, jornalero_id)
        if not row:
            return False
        await session.delete(row)
        await session.commit()
        return True


async def contar_jornaleros() -> int:
    async with async_session_factory() as session:
        return (await session.execute(select(func.count(JornaleroDb.id)))).scalar() or 0


# ── Seed desde Excel ───────────────────────────────────────────────────

async def seed_desde_excel(file_path: str | None = None) -> dict:
    import openpyxl

    path = file_path or _resolve_seed_path()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Archivo no encontrado: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Horas_Jornaleros" not in wb.sheetnames:
        wb.close()
        raise KeyError(f"La hoja 'Horas_Jornaleros' no existe en {path}. Hojas: {wb.sheetnames}")

    ws = wb["Horas_Jornaleros"]
    creados = 0
    actualizados = 0
    omitidos = 0

    async with async_session_factory() as session:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 13:
                omitidos += 1
                continue
            jornalero_id = str(row[0]).strip() if row[0] is not None else ""
            if not jornalero_id:
                omitidos += 1
                continue
            fecha_inicial = _parse_datetime_safe(row[1])
            fecha_final = _parse_datetime_safe(row[2])
            if not fecha_inicial or not fecha_final:
                omitidos += 1
                continue

            datos = {
                "fecha_inicial": fecha_inicial,
                "fecha_final": fecha_final,
                "tipo_trabajador": str(row[3]).strip() if row[3] else "JORNALERO",
                "cd": str(row[4]).strip() if row[4] else "",
                "unidad": str(row[5]).strip() if row[5] else "",
                "area": str(row[6]).strip() if row[6] else "DESPACHO",
                "cantidad_jornaleros": _parse_float_safe(row[7]),
                "horas_trabajadas": _parse_float_safe(row[8]),
                "dias_trabajados_totales": _parse_float_safe(row[9]),
                "dias_trabajados_laborales": _parse_float_safe(row[10]),
                "llenado_por": str(row[11]).strip() if row[11] else "",
                "fecha_creacion": _parse_datetime_safe(row[12]) or datetime.now(),
            }
            if not datos["cd"]:
                omitidos += 1
                continue

            existente = await session.get(JornaleroDb, jornalero_id)
            if existente:
                for key, val in datos.items():
                    if key == "fecha_creacion" and getattr(existente, key, None):
                        continue
                    setattr(existente, key, val)
                existente.estado_sincronizacion = "sincronizado"
                actualizados += 1
            else:
                row_db = JornaleroDb(
                    id=jornalero_id,
                    fecha_inicial=datos["fecha_inicial"],
                    fecha_final=datos["fecha_final"],
                    tipo_trabajador=datos["tipo_trabajador"],
                    cd=datos["cd"],
                    unidad=datos["unidad"],
                    area=datos["area"],
                    cantidad_jornaleros=datos["cantidad_jornaleros"],
                    horas_trabajadas=datos["horas_trabajadas"],
                    dias_trabajados_totales=datos["dias_trabajados_totales"],
                    dias_trabajados_laborales=datos["dias_trabajados_laborales"],
                    llenado_por=datos["llenado_por"],
                    fecha_creacion=datos["fecha_creacion"],
                    estado_sincronizacion="sincronizado",
                )
                session.add(row_db)
                creados += 1

        await session.commit()

    wb.close()
    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}


async def limpiar_todo():
    async with async_session_factory() as session:
        await session.execute(delete(JornaleroDb))
        await session.commit()
    logger.info("Tabla jornaleros limpiada")


# ── Pull desde Google Sheets (upsert por ID) ───────────────────────────

def _parse_sheet_row(headers: list[str], row: list) -> dict | None:
    """Convierte una fila de Sheets (lista) a dict usando los headers."""
    if not row or not row[0]:
        return None
    mapeo = {
        "ID": "id", "FECHA_INICIAL": "fecha_inicial", "FECHA_FINAL": "fecha_final",
        "TIPO_TRABAJADOR": "tipo_trabajador", "CD": "cd", "UNIDAD": "unidad",
        "AREA": "area", "CANTIDAD_JORNALEROS": "cantidad_jornaleros",
        "HORAS_TRABAJADAS": "horas_trabajadas", "DIAS_TRABAJADOS_TOTALES": "dias_trabajados_totales",
        "DIAS_TRABAJADOS_LABORALES": "dias_trabajados_laborales",
        "LLENADO POR": "llenado_por", "FECHA_CREACION": "fecha_creacion",
    }
    data: dict = {}
    for i, header in enumerate(headers):
        key = mapeo.get(str(header).strip().upper(), "")
        if not key:
            continue
        val = row[i] if i < len(row) else None
        if key in ("fecha_inicial", "fecha_final", "fecha_creacion"):
            data[key] = _parse_datetime_safe(val)
        elif key in ("cantidad_jornaleros", "horas_trabajadas", "dias_trabajados_totales", "dias_trabajados_laborales"):
            data[key] = _parse_float_safe(val)
        elif val is not None:
            data[key] = str(val).strip()
    if not data.get("id") or not data.get("fecha_inicial") or not data.get("fecha_final") or not data.get("cd"):
        return None
    return data


async def upsert_from_sheet_rows(rows: list[list]) -> dict:
    """Recibe filas crudas de Sheets (con headers) y hace upsert por ID en la BD."""
    if not rows:
        return {"creados": 0, "actualizados": 0, "omitidos": 0}
    headers = [str(h).strip().upper() for h in rows[0]]
    creados = 0
    actualizados = 0
    omitidos = 0

    async with async_session_factory() as session:
        for row in rows[1:]:
            datos = _parse_sheet_row(headers, row)
            if not datos:
                omitidos += 1
                continue
            existente = await session.get(JornaleroDb, datos["id"])
            if existente:
                for key, val in datos.items():
                    if key == "fecha_creacion" and val is None:
                        continue
                    setattr(existente, key, val)
                existente.estado_sincronizacion = "sincronizado"
                existente.error_sincronizacion = None
                actualizados += 1
            else:
                row_db = JornaleroDb(
                    id=datos["id"],
                    fecha_inicial=datos["fecha_inicial"],
                    fecha_final=datos["fecha_final"],
                    tipo_trabajador=datos.get("tipo_trabajador", "JORNALERO"),
                    cd=datos["cd"],
                    unidad=datos.get("unidad", ""),
                    area=datos.get("area", "DESPACHO"),
                    cantidad_jornaleros=datos.get("cantidad_jornaleros", 0.0),
                    horas_trabajadas=datos.get("horas_trabajadas", 0.0),
                    dias_trabajados_totales=datos.get("dias_trabajados_totales", 0.0),
                    dias_trabajados_laborales=datos.get("dias_trabajados_laborales", 0.0),
                    llenado_por=datos.get("llenado_por", ""),
                    fecha_creacion=datos.get("fecha_creacion") or datetime.now(),
                    estado_sincronizacion="sincronizado",
                )
                session.add(row_db)
                creados += 1

        await session.commit()

    return {"creados": creados, "actualizados": actualizados, "omitidos": omitidos}


async def obtener_pendientes() -> list[JornaleroDb]:
    async with async_session_factory() as session:
        stmt = select(JornaleroDb).where(JornaleroDb.estado_sincronizacion.in_(["pendiente", "error"]))
        result = await session.execute(stmt)
        return list(result.scalars().all())


async def obtener_todos() -> list[JornaleroDb]:
    async with async_session_factory() as session:
        result = await session.execute(select(JornaleroDb).order_by(JornaleroDb.fecha_inicial))
        return list(result.scalars().all())


async def marcar_sincronizado(jornalero_id: str, error: str | None = None):
    async with async_session_factory() as session:
        row = await session.get(JornaleroDb, jornalero_id)
        if row:
            row.estado_sincronizacion = "error" if error else "sincronizado"
            row.error_sincronizacion = error
            await session.commit()


# ── Estadísticas ───────────────────────────────────────────────────────

async def obtener_stats() -> dict:
    async with async_session_factory() as session:
        total = (await session.execute(select(func.count(JornaleroDb.id)))).scalar() or 0
        totales = (await session.execute(
            select(
                func.sum(JornaleroDb.cantidad_jornaleros),
                func.sum(JornaleroDb.horas_trabajadas),
                func.sum(JornaleroDb.dias_trabajados_totales),
                func.sum(JornaleroDb.dias_trabajados_laborales),
            )
        )).one()

        rows_cd = (await session.execute(
            select(
                JornaleroDb.cd,
                func.count(JornaleroDb.id),
                func.sum(JornaleroDb.cantidad_jornaleros),
                func.sum(JornaleroDb.horas_trabajadas),
            ).group_by(JornaleroDb.cd).order_by(JornaleroDb.cd)
        )).all()
        por_cd = [
            {
                "cd": r[0],
                "registros": r[1],
                "cantidad_jornaleros": round(r[2] or 0, 2),
                "horas_trabajadas": round(r[3] or 0, 2),
            }
            for r in rows_cd
        ]

        rows_area = (await session.execute(
            select(
                JornaleroDb.area,
                func.count(JornaleroDb.id),
                func.sum(JornaleroDb.cantidad_jornaleros),
                func.sum(JornaleroDb.horas_trabajadas),
            ).group_by(JornaleroDb.area).order_by(JornaleroDb.area)
        )).all()
        por_area = [
            {
                "area": r[0],
                "registros": r[1],
                "cantidad_jornaleros": round(r[2] or 0, 2),
                "horas_trabajadas": round(r[3] or 0, 2),
            }
            for r in rows_area
        ]

        pendientes = (await session.execute(
            select(func.count(JornaleroDb.id)).where(JornaleroDb.estado_sincronizacion.in_(["pendiente", "error"]))
        )).scalar() or 0

        return {
            "total_registros": total,
            "pendientes_sincronizacion": pendientes,
            "totales": {
                "cantidad_jornaleros": round(totales[0] or 0, 2),
                "horas_trabajadas": round(totales[1] or 0, 2),
                "dias_trabajados_totales": round(totales[2] or 0, 2),
                "dias_trabajados_laborales": round(totales[3] or 0, 2),
            },
            "por_cd": por_cd,
            "por_area": por_area,
        }
