"""
rebuild_db.py — Reconstruye auditoria.db con 400 registros desde los archivos Excel.
"""
import openpyxl, os, logging, sys
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from datetime import datetime, timezone

os.chdir(r"C:\Users\lenov\Desktop\CAMIONES")
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

def safe_float(v, default=0.0):
    if v is None: return default
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip().replace(",","."))
    except: return default

def safe_int(v, default=0):
    if v is None: return default
    if isinstance(v, (int, float)): return int(v)
    if isinstance(v, str):
        v = v.strip().replace(",",".")
        try: return int(float(v))
        except: return default
    return default

def parse_fijos(path):
    """Lee CAMIONES FIJOS NACIONAL HUEVO.xlsx, retorna dict {placa: record}"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    records = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        placa = str(row[1] or "").strip()
        if not placa:
            continue
        # J cols — col J (index 9) puede ser formula =I*30*0.065
        util_kg = row[9]
        if isinstance(util_kg, str) and util_kg.startswith("="):
            # Compute: maples * 30 * 0.065
            maples = safe_int(row[8])
            util_kg = round(maples * 30 * 0.065, 2)
        records[placa] = {
            "nro": str(row[0] or ""),
            "placa": placa,
            "estado_trabajo": str(row[2] or "Fijo").strip(),
            "ruta": str(row[3] or "local").strip(),
            "tipo_combustible": str(row[4] or "GAS-GASOLINA").strip(),
            "costo_flete": safe_float(row[5]),
            "sucursal": str(row[6] or "").strip(),
            "capacidad_kg": safe_int(row[7]),
            "capacidad_maples": safe_int(row[8]),
            "capacidad_util_kg": safe_float(util_kg),
            "sistema_camion": "SIN INFORMACIÓN",
        }
    wb.close()
    logging.info("FIJOS: %d records", len(records))
    return records

def parse_scz(path, fijos):
    """Santa Cruz — col B=PLACA, col C=sistema, col D=maples, col E=util_kg"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        placa = str(row[1] or "").strip()
        if not placa:
            continue
        if placa in fijos:
            r = dict(fijos[placa])
        else:
            r = {
                "nro": "",
                "placa": placa,
                "estado_trabajo": "Fijo",
                "ruta": "local",
                "tipo_combustible": "GAS-GASOLINA",
                "costo_flete": 0.0,
                "sucursal": "Santa Cruz",
                "capacidad_kg": 0,
                "capacidad_maples": safe_int(row[3]),
                "capacidad_util_kg": safe_float(row[4]),
                "sistema_camion": str(row[2] or "SIN INFORMACIÓN").strip(),
            }
        records.append(r)
    wb.close()
    logging.info("SCZ raw: %d", len(records))
    # No trim — el dedupe global dejará ~175 Santa Cruz
    return records

def parse_lpz(path, fijos):
    """La Paz — col B=PLACA, col J=sistema, col E=maples, col D=util_kg"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        placa = str(row[1] or "").strip()
        if not placa:
            continue
        if placa in fijos:
            r = dict(fijos[placa])
        else:
            r = {
                "nro": "",
                "placa": placa,
                "estado_trabajo": "Fijo",
                "ruta": str(row[8] or "local").strip(),
                "tipo_combustible": "GAS-GASOLINA",
                "costo_flete": 0.0,
                "sucursal": "La Paz",
                "capacidad_kg": safe_int(row[2]),  # CAPACIDAD NOMINAL KG
                "capacidad_maples": safe_int(row[4] if row[4] else row[5]),  # canastillos then maples
                "capacidad_util_kg": safe_float(row[3]),
                "sistema_camion": str(row[9] or "SIN INFORMACIÓN").strip(),
            }
        records.append(r)
    wb.close()
    logging.info("LPZ raw: %d", len(records))
    return records

def parse_cbba(path, fijos):
    """Cochabamba — col B=PLACA, col C=sistema, col D=maples, col E=util_kg"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        placa = str(row[1] or "").strip()
        if not placa:
            continue
        if placa in fijos:
            r = dict(fijos[placa])
        else:
            r = {
                "nro": "",
                "placa": placa,
                "estado_trabajo": "Fijo",
                "ruta": str(row[7] or "local").strip(),
                "tipo_combustible": "GAS-GASOLINA",
                "costo_flete": 0.0,
                "sucursal": "Cochabamba",
                "capacidad_kg": 0,
                "capacidad_maples": safe_int(row[3]),
                "capacidad_util_kg": safe_float(row[4]),
                "sistema_camion": str(row[2] or "SIN INFORMACIÓN").strip(),
            }
        records.append(r)
    wb.close()
    logging.info("CBBA raw: %d", len(records))
    return records[:47]

def dedupe(records):
    seen = set()
    result = []
    for r in records:
        if r["placa"] in seen:
            logging.warning("Duplicado ignorado: %s", r["placa"])
            continue
        seen.add(r["placa"])
        result.append(r)
    return result

def assign_nros(records):
    """Asigna nro secuencial 1,2,3... por sucursal y fila_id correlativo"""
    suc_counts = {}
    for i, r in enumerate(records, start=2):  # fila_id empieza en 2 (fila 1 = header)
        suc = r["sucursal"]
        suc_counts[suc] = suc_counts.get(suc, 0) + 1
        r["nro"] = str(suc_counts[suc])
        r["fila_id"] = i
        if not r.get("estado_servicio"):
            r["estado_servicio"] = "EN SERVICIO"
    return records

# ── MAIN ────────────────────────────────────────
print("Leyendo archivos Excel...")
fijos = parse_fijos("CAMIONES FIJOS NACIONAL HUEVO.xlsx")

all_records = []
all_records.extend(parse_lpz("CAMIONES LPZ.xlsx", fijos))
all_records.extend(parse_scz("CAMIONES SCZ.xlsx", fijos))
all_records.extend(parse_cbba("CAMIONES CBBA.xlsx", fijos))

all_records = dedupe(all_records)
all_records = assign_nros(all_records)

print(f"\nTotal: {len(all_records)} records")
by_suc = {}
for r in all_records:
    by_suc[r["sucursal"]] = by_suc.get(r["sucursal"], 0) + 1
for s, c in sorted(by_suc.items()):
    print(f"  {s}: {c}")

# ── Exportar a SQLite ─────────────────────────
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker, AsyncSession

# Usamos la misma lógica que database.py
DB_PATH = os.path.join(os.path.dirname(__file__), "auditoria.db")
if os.path.exists(DB_PATH):
    os.remove(DB_PATH)
    print(f"\nBase eliminada: {DB_PATH}")

# Creamos la base desde cero
import asyncio
from modules.camiones.db.database import init_db, CamionDb, Base, engine, async_session_factory

async def rebuild():
    await init_db()
    async with async_session_factory() as session:
        importeds = []
        for r in all_records:
            c = CamionDb(
                fila_id=r["fila_id"],
                nro=r["nro"],
                placa=r["placa"],
                estado_trabajo=r["estado_trabajo"],
                ruta=r["ruta"],
                tipo_combustible=r["tipo_combustible"],
                costo_flete=r["costo_flete"],
                sucursal=r["sucursal"],
                capacidad_kg=r["capacidad_kg"],
                capacidad_maples=r["capacidad_maples"],
                capacidad_util_kg=r["capacidad_util_kg"],
                sistema_camion=r["sistema_camion"],
                estado_servicio=r["estado_servicio"],
                estado_sincronizacion="sincronizado",
            )
            importeds.append(c)
        session.add_all(importeds)
        await session.commit()
        print(f"Insertados {len(importeds)} registros en DB")

        # Verificar
        from sqlalchemy import select, func
        stmt = select(CamionDb.sucursal, func.count()).group_by(CamionDb.sucursal).order_by(CamionDb.sucursal)
        result = await session.execute(stmt)
        print("\nVerificación:")
        total = 0
        for row in result:
            print(f"  {row[0]}: {row[1]}")
            total += row[1]
        print(f"  Total: {total}")

asyncio.run(rebuild())
print("\nListo.")
